# -*- coding: utf-8 -*-
"""
layout_premium_mc4.py
Versao: 1.3.28-producao-forense

Camada de layout POS-PROCESSAMENTO para o Motor Fiscal RIR70 by MC4.

Arquitetura:
- Motor gera Excel técnico neutro.
- Layout aplica apresentação visual final.
- Modo padrão é rápido para produção com muitos XMLs e Excel grande.
- Modo auditoria preserva mais evidência e aplica tratamento mais profundo.

Modos:
- rapido: cabeçalho, largura, filtro, congelamento, grade removida, abas coloridas. Não varre célula a célula em abas grandes.
- executivo: premium em abas pequenas; rápido em abas grandes.
- auditoria: limpeza profunda, sidecar completo e RAW preservado pelo motor.

Contrato de segurança:
- Não altera valores calculados.
- Não altera fórmulas.
- Não altera nomes de abas.
- Não remove colunas.
- Não cria tabelas estruturadas.
- Só altera apresentação visual e cabeçalho visível.
"""

from __future__ import annotations

from pathlib import Path
from datetime import datetime
from typing import Any, Dict, Optional
import hashlib
import json
import re
import shutil
import time
import unicodedata

try:
    from copy import copy
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
except Exception as exc:
    raise RuntimeError("openpyxl nao instalado. Execute: pip install openpyxl") from exc


PALETA = {
    "navy": "102A43",
    "blue": "1F4E79",
    "teal": "0F766E",
    "green": "548235",
    "amber": "C55A11",
    "rose": "C00000",
    "violet": "5B2C83",
    "slate": "334155",
    "white": "FFFFFF",
    "zebra": "F7F9FC",
    "line": "E5E7EB",
    "text": "111827",
    "muted": "475569",
    "ok_bg": "E2EFDA",
    "warn_bg": "FFF2CC",
    "err_bg": "FCE4D6",
}

TAB_COLORS = {
    "CAPA": PALETA["navy"],
    "FONTES_ANALISADAS": PALETA["blue"],
    "INCONSISTENCIAS": PALETA["rose"],
    "01_MOVIMENTO_VALIDADO": PALETA["blue"],
    "02_ARBITRAMENTO": PALETA["teal"],
    "03_INVENTARIO_VALORIZADO": PALETA["green"],
    "CUSTO_CMV": PALETA["teal"],
    "04_PENDENCIAS": PALETA["rose"],
    "05_FONTES_PROCESSADAS": PALETA["violet"],
    "AJUSTES_ANALISTA": PALETA["amber"],
    "AJUSTES": PALETA["amber"],
    "INSTRUCOES": PALETA["slate"],
}
HEADER_COLORS = dict(TAB_COLORS)

DEFAULT_CONFIG = {
    "layout_nivel": "executivo",
    "max_largura_coluna": 42,
    "modo_rapido_acima_linhas": 10000,
    "ordenar_chave_nfe_pre_validacao": True,
    "ordenar_chave_nfe_movimento": False,
    "limpar_estilos_residuais_apos_dados": True,
    "linhas_residuais_para_limpar": 80,
    "altura_linha_corpo": 18,
    "altura_linha_cabecalho": 34,
    "usar_zebra": True,
    "aplicar_zebra": "Sim",
    "estilizar_status": True,
    "validacao_profunda": "Nao",
    "fonte_corpo": "Aptos Narrow",
    "fonte_cabecalho": "Aptos Narrow",
    "tamanho_corpo": 10,
    "tamanho_cabecalho": 10,
    "usar_cabecalhos_amigaveis": True,
}

HEADER_FRIENDLY_NAMES = {
    "DESCRICAO_TECNICA": "Descrição Técnica",

    "VOLUME_ML": "Volume Unitário (ml)",
    "VOLUME_UNITARIO_ML": "Volume Unitário (ml)",
    "VOLUME_TOTAL_EMBALAGEM_ML": "Volume Total Embalagem (ml)",
    "FATOR_UNIDADE": "Fator Qtd Embalagem (un)",
    "FATOR_UNIDADE_MATRIZ": "Fator Qtd Embalagem Matriz (un)",
    "FATOR_QTD_EMBALAGEM_UN": "Fator Qtd Embalagem (un)",
    "UNIDADE_MEDIDA": "Unidade Medida Embalagem",
    "UNIDADE_MEDIDA_EMBALAGEM": "Unidade Medida Embalagem",
    "PRECO_MEDIO": "Preço Médio",
    "VARIACAO_PRECO (%)": "Variação Preço (%)",
    "DOCUMENTO_MAIOR_PRECO": "Documento Maior Preço",
    "EMPRESA_ORIGEM_PRECO": "Empresa Origem Preço",
    "PRODUTO_ORIGEM_PRECO": "Produto Origem Preço",
    "MES_ORIGEM_PRECO": "Mês Origem Preço",
    "JUSTIFICATIVA_PRECO_REF": "Justificativa Preço Referência",
    "EVIDENCIA_OBRIGATORIA": "Evidência Obrigatória",
    "RESPONSAVEL_DECISAO": "Responsável Decisão",
    "DECISAO_ANALISTA": "Decisão Analista",
    "SITUACAO": "Situação",
    "PERIODO": "Período",
    "CODIGO_CONTROLE_ARQUIVO": "Código Controle Arquivo",

    "INDICADOR": "Indicador",
    "VALOR": "Valor",
    "ARQUIVO": "Arquivo",
    "SHEET": "Aba",
    "TIPO_FONTE": "Tipo Fonte",
    "STATUS_PRE_VALIDACAO": "Status Pré-Validação",
    "DETALHE": "Detalhe",
    "CHAVE_NFE": "Chave NF-e",
    "DATA_EMISSAO": "Data Emissão",
    "ITENS": "Itens",
    "SHA256": "Hash Arquivo",
    "SHA256_CONTAINER": "Hash Pacote",
    "REGISTROS_LIDOS": "Registros Lidos",
    "REGISTROS_VALIDOS": "Registros Válidos",
    "REGISTROS_DUPLICADOS": "Registros Duplicados",
    "REGISTROS_BLOQUEADOS_CFOP": "Registros Bloqueados CFOP",
    "STATUS": "Status",
    "EMPRESA": "Empresa",
    "MES": "Mês",
    "COMPETENCIA": "Competência",
    "NUMERO_NF": "Número NF",
    "SERIE": "Série",
    "CODIGO_PRODUTO": "Código Produto",
    "CODIGO_ITEM": "Código Item",
    "DESCRICAO_PRODUTO": "Descrição Produto",
    "DESCRICAO": "Descrição",
    "NCM": "NCM",
    "CEST": "CEST",
    "CFOP": "CFOP",
    "QTD_CALCULO": "Quantidade Cálculo",
    "QTDE_INVENTARIO": "Quantidade Inventário",
    "VALOR_PRODUTO_BASE": "Valor Produto Base",
    "VALOR_COMERCIAL": "Valor Comercial",
    "VALOR_TRIBUTAVEL": "Valor Tributável",
    "VALOR_ESTOQUE_ARBITRADO": "Valor Estoque Arbitrado",
    "PARTICIPA_CALCULO": "Participa Cálculo",
    "MOTIVO_CLASSIFICACAO": "Motivo Classificação",
    "MOTIVO_EXCLUSAO": "Motivo Exclusão",
    "MAIOR_PRECO_UNITARIO": "Maior Preço Unitário",
    "MAIOR_VLR_UNITARIO": "Maior Vlr Unitário",
    "VLR_UNITARIO_BASE_RIR70": "Vlr Unitário Base RIR70",
    "FATOR_ARBITRAMENTO": "Fator Arbitramento",
    "BASE_ARBITRADA_70": "Base Arbitrada 70%",
    "CUSTO_ARBITRADO_70": "Custo Arbitrado 70%",
    "CMV_COMPETENCIA": "CMV Competência",
    "FONTE_NORMA": "Fonte Norma",
    "FORMULA_APLICADA": "Fórmula Aplicada",
    "TRAIL_CALCULO": "Memória de Cálculo",
    "VALORES_INTERMEDIARIOS": "Valores Intermediários",
    "ARREDONDAMENTO_APLICADO": "Arredondamento Aplicado",
    "HASH_LINHA_CALCULO": "Hash Linha Cálculo",
    "ID_REGRA_FISCAL": "ID Regra Fiscal",
    "HASH_FONTE": "Hash Fonte",
    "VERSAO_FONTE": "Versão Fonte",
    "FONTE_PREMISSA_OPERACIONAL": "Fonte Premissa Operacional",
    "FORMULA_CMV": "Fórmula CMV",
    "FORMULA_ESTOQUE": "Fórmula Estoque",
    "VALORES_INTERMEDIARIOS_ESTOQUE": "Valores Intermediários Estoque",
    "ID_CUSTO_CMV": "ID Custo CMV",
    "ID_INVENTARIO": "ID Inventário",
    "JUSTIFICATIVA_CRITERIO_SUBSIDIARIO": "Justificativa Critério Subsidiário",
    "INTERVENCAO_ANALISTA": "Intervenção Analista",
    "ACAO_ANALISTA": "Ação Analista",
    "TIPO_PENDENCIA": "Tipo Pendência",
    "NIVEL": "Nível",
    "ALTERA_CALCULO": "Altera Cálculo",
    "ID_PENDENCIA": "ID Pendência",
    "TIPO_AJUSTE": "Tipo Ajuste",
    "STATUS_AJUSTE": "Status Ajuste",
    "VALOR_REFERENCIA_INFORMADO": "Valor Referência Informado",
    "CAMINHO_EVIDENCIA": "Caminho Evidência",
    "JUSTIFICATIVA_COMPARABILIDADE": "Justificativa Comparabilidade",
    "RESPONSAVEL_VALIDACAO_FISCAL": "Responsável Validação Fiscal",
    "DATA_DECISAO": "Data Decisão",
    "TIPO_PRODUTO": "Tipo Produto",
    "CATEGORIA_ITEM": "Categoria Item",
    "CRITERIO_CUSTEIO_RIR70": "Critério Custeio RIR70",
    "ORIGEM_DO_PRECO": "Origem do Preço",
    "SITUACAO_PRECO": "Situação Preço",
    "ADVERTENCIA": "Advertência",
    "ALERTA": "Alerta",
}


def _as_bool(v: Any, default: bool = False) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return default
    s = str(v).strip().upper()
    if s in {"SIM", "S", "TRUE", "1", "YES", "Y"}:
        return True
    if s in {"NAO", "NÃO", "N", "FALSE", "0", "NO"}:
        return False
    return default


def _int(v: Any, default: int) -> int:
    try:
        return int(v)
    except Exception:
        return default


def _norm_header(value: Any) -> str:
    t = "" if value is None else str(value).strip()
    t = unicodedata.normalize("NFKD", t).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Z0-9]+", "_", t.upper()).strip("_")


def _title_from_header(value: Any) -> str:
    raw = "" if value is None else str(value).strip()
    key = _norm_header(raw)
    if key in HEADER_FRIENDLY_NAMES:
        return HEADER_FRIENDLY_NAMES[key]
    words = [w for w in re.split(r"[_\s]+", raw) if w]
    keep_upper = {"NF", "NFE", "NF-E", "NCM", "CEST", "CFOP", "ICMS", "IPI", "PIS", "COFINS", "ST", "FCP", "SKU", "CNPJ", "CPF", "UF", "ID", "XML", "ZIP", "RIR70", "CMV"}
    out = []
    for w in words:
        wu = w.upper()
        out.append(wu if wu in keep_upper else w.lower().capitalize())
    return " ".join(out).replace("Nfe", "NF-e")


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with Path(path).open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _cell_has_value(cell) -> bool:
    return cell.value is not None and str(cell.value).strip() != ""


def _true_max_col_by_header(ws) -> int:
    last = 1
    for c in range(1, ws.max_column + 1):
        if _cell_has_value(ws.cell(1, c)):
            last = c
    return last


def _true_last_row_by_values(ws, max_col: int) -> int:
    last = 1
    for r in range(1, ws.max_row + 1):
        for c in range(1, max_col + 1):
            if _cell_has_value(ws.cell(r, c)):
                last = r
                break
    return last


def _technical_headers(ws, max_col: int) -> list[str]:
    return [_norm_header(ws.cell(1, c).value) for c in range(1, max_col + 1)]


def _clear_column_dimensions(ws):
    try:
        ws.column_dimensions.clear()
    except Exception:
        try:
            for key in list(ws.column_dimensions.keys()):
                del ws.column_dimensions[key]
        except Exception:
            pass


def _reset_cell_visual(cell, cfg: Dict[str, Any]):
    cell.fill = PatternFill(fill_type=None)
    cell.border = Border()
    cell.font = Font(name=cfg["fonte_corpo"], size=cfg["tamanho_corpo"], color=PALETA["text"])
    cell.alignment = Alignment(horizontal="general", vertical="bottom", wrap_text=False)
    cell.protection = copy(Protection())


def _clean_residual_styles(ws, true_last_row: int, true_max_col: int, cfg: Dict[str, Any], deep: bool):
    if not _as_bool(cfg.get("limpar_estilos_residuais_apos_dados"), True):
        return
    rows_to_clean = _int(cfg.get("linhas_residuais_para_limpar"), 80 if not deep else 400)
    max_row = min(ws.max_row, true_last_row + rows_to_clean)

    for r in range(true_last_row + 1, max_row + 1):
        ws.row_dimensions[r].height = None
        for c in range(1, max(ws.max_column, true_max_col) + 1):
            cell = ws.cell(r, c)
            if not _cell_has_value(cell):
                _reset_cell_visual(cell, cfg)

    for c in range(true_max_col + 1, ws.max_column + 1):
        col_letter = get_column_letter(c)
        try:
            ws.column_dimensions[col_letter].hidden = True
            ws.column_dimensions[col_letter].width = 2
        except Exception:
            pass
        if deep:
            for r in range(1, min(ws.max_row, true_last_row + rows_to_clean) + 1):
                cell = ws.cell(r, c)
                if not _cell_has_value(cell):
                    _reset_cell_visual(cell, cfg)


def _sort_by_chave_nfe(ws, true_max_col: int) -> bool:
    hdr = _technical_headers(ws, true_max_col)
    if "CHAVE_NFE" not in hdr or ws.max_row <= 2:
        return False
    col = hdr.index("CHAVE_NFE") + 1
    true_last_row = _true_last_row_by_values(ws, true_max_col)

    def key(row_values):
        val = row_values[col - 1] if col - 1 < len(row_values) else ""
        s = str(val or "").strip()
        d = re.sub(r"\D", "", s)
        if d:
            try:
                return (0, int(d), s)
            except Exception:
                return (0, d, s)
        return (1, "", s)

    rows = [list(r) for r in ws.iter_rows(min_row=2, max_row=true_last_row, max_col=true_max_col, values_only=True)]
    for r_idx, row in enumerate(sorted(rows, key=key), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = value
    return True


def _apply_friendly_headers(ws, true_max_col: int, cfg: Dict[str, Any]) -> Dict[str, str]:
    if not _as_bool(cfg.get("usar_cabecalhos_amigaveis"), True):
        return {}
    mapping: Dict[str, str] = {}
    for c in range(1, true_max_col + 1):
        cell = ws.cell(1, c)
        original = "" if cell.value is None else str(cell.value).strip()
        friendly = _title_from_header(original)
        mapping[original] = friendly
        cell.value = friendly
    return mapping


def _column_width(header_display: str, header_key: str, max_len: int, max_width: int) -> int:
    """Largura enxuta e previsível.

    Não deixa código/produto virar coluna gigante. Textos longos ficam com quebra,
    e evidências técnicas são legíveis sem inflar a planilha.
    """
    h = header_key.upper().strip()
    special = {
        "ID_INVENTARIO": 18, "ID_CUSTO_CMV": 18, "ID_ARBITRAMENTO": 18,
        "EMPRESA": 18, "PERFIL_EMPRESA": 16, "MES": 12, "ALMOX": 10,
        "CODIGO_PRODUTO": 18, "CODIGO_ITEM": 18, "DOCUMENTO": 14,
        "NCM": 10, "CEST": 12, "CFOP": 8, "UNIDADE": 10,
        "FATOR_QTD_EMBALAGEM_UN": 18, "FATOR_UNIDADE_MATRIZ": 18,
        "VOLUME_UNITARIO_ML": 18, "VOLUME_TOTAL_EMBALAGEM_ML": 22,
        "TIPO_PRODUTO": 18, "PART_ARBITRAMENTO": 16, "VINCULO_GRUPO": 14,
        "CATEGORIA_ITEM": 16, "CRITERIO_CUSTEIO_RIR70": 22,
        "DESCRICAO_PRODUTO": 34, "DESCRICAO_TECNICA": 34, "DESCRICAO": 34,
        "MAIOR_PRECO_UNITARIO": 18, "CUSTO_ARBITRADO_70": 18, "PRECO_MEDIO": 16,
        "QUANTIDADE_VENDIDA": 18, "CMV": 16, "SALDO_ESTOQUE": 16,
        "VALOR_ESTOQUE_ARBITRADO": 20, "VALOR_PRODUTO": 18, "VALOR_UNITARIO": 18,
        "ORIGEM_DO_PRECO": 22, "EMPRESA_ORIGEM_PRECO": 18, "PRODUTO_ORIGEM_PRECO": 20,
        "MES_ORIGEM_PRECO": 14, "DOCUMENTO_ORIGEM": 20,
        "SITUACAO_PRECO": 16, "ALERTA": 14, "ADVERTENCIA": 26,
        "INTERVENCAO_ANALISTA": 18, "ACAO_ANALISTA": 34,
        "FONTE_NORMA": 34, "FONTE_PREMISSA_OPERACIONAL": 38,
        "FORMULA_APLICADA": 34, "FORMULA_CMV": 30, "FORMULA_ESTOQUE": 34,
        "TRAIL_CALCULO": 38, "VALORES_INTERMEDIARIOS": 38, "VALORES_INTERMEDIARIOS_ESTOQUE": 38,
        "ARREDONDAMENTO_APLICADO": 34, "ID_REGRA_FISCAL": 24, "HASH_LINHA_CALCULO": 22,
        "HASH_FONTE": 22, "CODIGO_CONTROLE_ARQUIVO": 22,
        "CHAVE_NFE": 30, "ARQUIVO": 34, "CAMINHO_EVIDENCIA": 34,
        "TIPO_PENDENCIA": 28, "NIVEL": 14, "RISCO_SE_NAO_CORRIGIR": 38,
        "MOTIVO_PRINCIPAL_DESCARTE": 34,
    }
    if h in special:
        return min(max_width, special[h])
    if any(x in h for x in ("HASH", "SHA256")):
        return min(max_width, 22)
    if any(x in h for x in ("JUSTIFICATIVA", "MOTIVO", "DETALHE", "EVIDENCIA", "FONTE", "TRAIL", "VALORES_INTERMEDIARIOS", "ACAO")):
        return min(max_width, 36)
    if any(x in h for x in ("DESCRICAO", "PRODUTO")):
        return min(max_width, 34)
    if any(x in h for x in ("STATUS", "SITUACAO", "NIVEL")):
        return min(max_width, 18)
    if "DATA" in h:
        return 14
    if any(x in h for x in ("QTD", "QTDE", "QUANTIDADE")):
        return 16
    return max(9, min(max(max_len + 2, len(header_display) + 3), max_width))


def _status_fill(value: Any):
    s = str(value or "").upper()
    if any(x in s for x in ("ERRO", "INVALID", "INVÁLID", "MALFORM", "BLOQUE", "CRITICO", "CRÍTICO", "CANCEL", "SEM_XML")):
        return PatternFill("solid", fgColor=PALETA["err_bg"])
    if any(x in s for x in ("REVIS", "ATEN", "PEND", "SEM_PROTOCOLO", "ALERTA")):
        return PatternFill("solid", fgColor=PALETA["warn_bg"])
    if any(x in s for x in ("OK", "VALID", "AUTORIZ", "APROV")):
        return PatternFill("solid", fgColor=PALETA["ok_bg"])
    return None


def _mode_settings(cfg: Dict[str, Any], rows: int) -> Dict[str, Any]:
    nivel = str(cfg.get("layout_nivel") or "rapido").strip().lower()
    threshold = _int(cfg.get("modo_rapido_acima_linhas"), 10000)
    is_large = rows > threshold

    if nivel == "auditoria":
        return {"nivel": nivel, "deep": True, "body_style_rows": min(rows, 50000), "zebra": True, "status": True}
    if nivel == "executivo":
        if is_large:
            return {"nivel": "executivo-rapido", "deep": False, "body_style_rows": min(rows, 5000), "zebra": False, "status": False}
        return {"nivel": nivel, "deep": True, "body_style_rows": rows, "zebra": True, "status": True}

    # produção rápida: só faz o essencial; evita varredura célula a célula em grandes abas.
    return {
        "nivel": "rapido",
        "deep": False,
        "body_style_rows": min(rows, 1000),
        "zebra": _as_bool(cfg.get("aplicar_zebra", cfg.get("usar_zebra")), False) and not is_large,
        "status": False if is_large else _as_bool(cfg.get("estilizar_status"), True),
    }



def _aplicar_ajustes_capa_executivo_auditoria(ws, cfg: Dict[str, Any], mode: Dict[str, Any]) -> Dict[str, Any]:
    """Ajustes microcirurgicos da aba CAPA solicitados para modo executivo/auditoria.

    Escopo:
    - Somente aba CAPA.
    - Somente quando layout_nivel for executivo ou auditoria.
    - Nao altera valores, formulas, nomes de abas ou calculos.
    """
    nivel = str(cfg.get("layout_nivel") or "").strip().lower()
    modo_calculado = str(mode.get("nivel") or "").strip().lower()
    aplicar = nivel in {"executivo", "auditoria"} or modo_calculado.startswith("executivo") or modo_calculado == "auditoria"
    if ws.title.upper() != "CAPA" or not aplicar:
        return {"aplicado": False}

    alteracoes = []

    # B2 e B17: alinhamento no meio.
    for ref in ("B2", "B17"):
        cell = ws[ref]
        current = cell.alignment
        cell.alignment = Alignment(
            horizontal=current.horizontal or "left",
            vertical="center",
            text_rotation=current.text_rotation,
            wrap_text=current.wrap_text,
            shrink_to_fit=current.shrink_to_fit,
            indent=current.indent,
        )
        alteracoes.append(f"{ref}: alinhamento vertical centro")

    # B9, B18, B70: quebrar texto automaticamente.
    for ref in ("B9", "B18", "B70"):
        cell = ws[ref]
        current = cell.alignment
        cell.alignment = Alignment(
            horizontal=current.horizontal or "left",
            vertical=current.vertical or "top",
            text_rotation=current.text_rotation,
            wrap_text=True,
            shrink_to_fit=current.shrink_to_fit,
            indent=current.indent,
        )
        alteracoes.append(f"{ref}: quebrar texto automaticamente")

    # B65:B68: retirar negrito.
    for row in range(65, 69):
        cell = ws[f"B{row}"]
        f = cell.font
        cell.font = Font(
            name=f.name or cfg.get("fonte_corpo", "Aptos Narrow"),
            sz=f.sz or cfg.get("tamanho_corpo", 10),
            b=False,
            i=f.i,
            u=f.u,
            strike=f.strike,
            color=f.color,
            vertAlign=f.vertAlign,
            charset=f.charset,
            family=f.family,
            scheme=f.scheme,
            outline=f.outline,
            shadow=f.shadow,
            condense=f.condense,
            extend=f.extend,
        )
        alteracoes.append(f"B{row}: remover negrito")

    return {"aplicado": True, "alteracoes": alteracoes}




def _sentence_case_primeira_letra(valor: Any) -> Any:
    """Coloca a primeira letra alfabetica em maiuscula sem rebaixar o restante."""
    if valor is None:
        return valor
    s = str(valor)
    for i, ch in enumerate(s):
        if ch.isalpha():
            return s[:i] + ch.upper() + s[i+1:]
    return valor


def _title_case_simples(valor: Any) -> Any:
    """Primeira letra de cada palavra em maiuscula, preservando valores vazios."""
    if valor is None:
        return valor
    s = str(valor).strip()
    if not s:
        return valor
    return " ".join(p[:1].upper() + p[1:].lower() if p else p for p in s.split())


def _normalizar_trail_motivo(valor: Any) -> Any:
    """Depois da palavra motivo, coloca a primeira letra seguinte em maiuscula."""
    if valor is None:
        return valor
    s = str(valor)

    def repl(m):
        return m.group(1) + m.group(2).upper()

    return re.sub(r"(?i)(\bmotivo\b\s*[:\-]?\s*)([a-záàâãéêíóôõúç])", repl, s)


def _substituir_inferido(valor: Any) -> Any:
    """Troca 'inferido' por termo mais comum, sem alterar demais o texto."""
    if valor is None:
        return valor
    s = str(valor)
    return re.sub(r"(?i)\binferido\b", "Estimado", s)


def _font_with_color(base_font, color="111827", bold=None):
    """Copia a fonte existente com troca controlada de cor/negrito."""
    return Font(
        name=base_font.name,
        sz=base_font.sz,
        b=base_font.b if bold is None else bold,
        i=base_font.i,
        u=base_font.u,
        strike=base_font.strike,
        color=color,
        vertAlign=base_font.vertAlign,
        charset=base_font.charset,
        family=base_font.family,
        scheme=base_font.scheme,
        outline=base_font.outline,
        shadow=base_font.shadow,
        condense=base_font.condense,
        extend=base_font.extend,
    )


def _sem_preenchimento(cell):
    cell.fill = PatternFill(fill_type=None)


def _aplicar_ajustes_movimento_validado_executivo_auditoria(ws, cfg: Dict[str, Any], mode: Dict[str, Any]) -> Dict[str, Any]:
    """Ajustes microcirurgicos da aba MOVIMENTO_VALIDADO.

    Escopo:
    - Somente 01_MOVIMENTO_VALIDADO ou MOVIMENTO_VALIDADO.
    - Somente layout executivo/auditoria.
    - Nao remove colunas, nao mexe em calculos, nao altera formulas.
    """
    nivel = str(cfg.get("layout_nivel") or "").strip().lower()
    modo_calculado = str(mode.get("nivel") or "").strip().lower()
    aplicar = nivel in {"executivo", "auditoria"} or modo_calculado.startswith("executivo") or modo_calculado == "auditoria"
    nome = ws.title.upper().strip()
    if nome not in {"01_MOVIMENTO_VALIDADO", "MOVIMENTO_VALIDADO"} or not aplicar:
        return {"aplicado": False}

    true_max_col = _true_max_col_by_header(ws)
    true_last_row = _true_last_row_by_values(ws, true_max_col)
    if true_last_row < 2:
        return {"aplicado": True, "linhas": true_last_row, "observacao": "sem linhas de dados"}

    alteracoes = []

    # Alinhamento meio no cabecalho/area principal da aba, sem varrer colunas fora da area real.
    for c in range(1, true_max_col + 1):
        cell = ws.cell(1, c)
        cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "center", vertical="center", wrap_text=False)
    alteracoes.append("cabecalho: alinhamento vertical centro")

    RED = "C00000"
    BLACK = "000000"

    # Coluna T - FATOR_CAIXA: centralizar.
    for r in range(2, true_last_row + 1):
        cell = ws[f"T{r}"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    alteracoes.append("T/FATOR_CAIXA: centralizar")

    # Coluna W - SITUACAO_QTD: OK -> Validado; demais em vermelho; primeira letra maiuscula.
    for r in range(2, true_last_row + 1):
        cell = ws[f"W{r}"]
        val = "" if cell.value is None else str(cell.value).strip()
        if val.upper() == "OK":
            cell.value = "Validado"
            cell.font = _font_with_color(cell.font, BLACK)
        else:
            if val:
                cell.value = _title_case_simples(val)
                if str(cell.value).strip() != "Validado":
                    cell.font = _font_with_color(cell.font, RED)
            else:
                cell.font = _font_with_color(cell.font, BLACK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    alteracoes.append("W/SITUACAO_QTD: OK renomeado para Validado; demais em vermelho")

    # Coluna X - CALCULA: retirar fundo abaixo do cabecalho; manter tudo maiusculo; NAO em vermelho.
    for r in range(2, true_last_row + 1):
        cell = ws[f"X{r}"]
        _sem_preenchimento(cell)
        val = "" if cell.value is None else str(cell.value).strip().upper()
        cell.value = val if val else cell.value
        if val == "NAO":
            cell.font = _font_with_color(cell.font, RED)
        else:
            cell.font = _font_with_color(cell.font, BLACK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    alteracoes.append("X/CALCULA: sem fundo; NAO em vermelho; valores em maiusculo")

    # Coluna Y - MOTIVO: manter por rastreabilidade, pois nao e igual a ADVERTENCIA.
    alteracoes.append("Y/MOTIVO: mantida por rastreabilidade; nao removida")

    # Coluna Z - ALERTA: retirar fundo abaixo do cabecalho; manter cores das letras.
    for r in range(2, true_last_row + 1):
        cell = ws[f"Z{r}"]
        _sem_preenchimento(cell)
        cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "left", vertical="center", wrap_text=False)
    alteracoes.append("Z/ALERTA: sem fundo, mantendo cor das letras")

    # Coluna AA - ADVERTENCIA: letras pretas, sem fundo, inferido -> Estimado.
    for r in range(2, true_last_row + 1):
        cell = ws[f"AA{r}"]
        cell.value = _substituir_inferido(cell.value)
        _sem_preenchimento(cell)
        cell.font = _font_with_color(cell.font, BLACK)
        cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "left", vertical="center", wrap_text=True)
    alteracoes.append("AA/ADVERTENCIA: sem fundo, letra preta, Inferido -> Estimado")

    # Coluna AB - ID_REGRA_FISCAL: primeira letra maiuscula, de forma conservadora.
    for r in range(2, true_last_row + 1):
        cell = ws[f"AB{r}"]
        cell.value = _sentence_case_primeira_letra(cell.value)
        cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "left", vertical="center", wrap_text=False)
    alteracoes.append("AB/ID_REGRA_FISCAL: primeira letra alfabetica em maiuscula")

    # Coluna AD - FORMULA_APLICADA: primeira letra maiuscula.
    for r in range(2, true_last_row + 1):
        cell = ws[f"AD{r}"]
        cell.value = _sentence_case_primeira_letra(cell.value)
        cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "left", vertical="center", wrap_text=True)
    alteracoes.append("AD/FORMULA_APLICADA: primeira letra alfabetica em maiuscula")

    # Coluna AE - TRAIL_CALCULO: apos a palavra motivo, primeira letra maiuscula.
    for r in range(2, true_last_row + 1):
        cell = ws[f"AE{r}"]
        cell.value = _normalizar_trail_motivo(cell.value)
        cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "left", vertical="top", wrap_text=True)
    alteracoes.append("AE/TRAIL_CALCULO: primeira letra apos 'motivo' em maiuscula")

    # Coluna AH - INTERVENCAO_ANALISTA: centralizar; SIM em vermelho.
    for r in range(2, true_last_row + 1):
        cell = ws[f"AH{r}"]
        val = "" if cell.value is None else str(cell.value).strip().upper()
        if val:
            cell.value = val
        if val == "SIM":
            cell.font = _font_with_color(cell.font, RED)
        else:
            cell.font = _font_with_color(cell.font, BLACK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    alteracoes.append("AH/INTERVENCAO_ANALISTA: centralizar; SIM em vermelho")

    return {
        "aplicado": True,
        "linhas_processadas": true_last_row - 1,
        "observacao": "MOTIVO mantida por diferenciar causa/rastreabilidade de ADVERTENCIA.",
        "alteracoes": alteracoes,
    }




def _parse_percentual_mc4(valor: Any):
    """Converte valores percentuais em float percentual; retorna None se nao interpretar."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        # Excel pode armazenar 0.85 como 85% ou 85 como 85.
        return float(valor) * 100 if abs(float(valor)) <= 1.5 else float(valor)
    s = str(valor).strip()
    if not s:
        return None
    s = s.replace("%", "").replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _col_index_por_header_mc4(ws, nomes):
    """Busca coluna por nomes tecnicos ou visuais no cabecalho."""
    alvo = {_norm_header(n) for n in nomes}
    max_col = _true_max_col_by_header(ws)
    for c in range(1, max_col + 1):
        if _norm_header(ws.cell(1, c).value) in alvo:
            return c
    return None


def _aplicar_ajustes_arbitramento_executivo_auditoria(ws, cfg: Dict[str, Any], mode: Dict[str, Any]) -> Dict[str, Any]:
    """Ajustes microcirurgicos da aba ARBITRAMENTO.

    Escopo:
    - Somente 02_ARBITRAMENTO ou ARBITRAMENTO.
    - Somente layout executivo/auditoria.
    - Nao altera formulas, regras fiscais, calculos ou nomes de abas.
    """
    nivel = str(cfg.get("layout_nivel") or "").strip().lower()
    modo_calculado = str(mode.get("nivel") or "").strip().lower()
    aplicar = nivel in {"executivo", "auditoria"} or modo_calculado.startswith("executivo") or modo_calculado == "auditoria"
    nome = ws.title.upper().strip()
    if nome not in {"02_ARBITRAMENTO", "ARBITRAMENTO"} or not aplicar:
        return {"aplicado": False}

    true_max_col = _true_max_col_by_header(ws)
    true_last_row = _true_last_row_by_values(ws, true_max_col)
    if true_last_row < 2:
        return {"aplicado": True, "linhas": true_last_row, "observacao": "sem linhas de dados"}

    alteracoes = []
    RED = "C00000"
    GREEN = "548235"
    GOLD = "C9A227"
    BLACK = "000000"

    # Localizacao por cabecalho, com fallback pelas letras citadas.
    col_custo = _col_index_por_header_mc4(ws, ["CUSTO_ARBITRADO_70", "Custo Arbitrado 70%"])
    col_variacao = _col_index_por_header_mc4(ws, ["VARIACAO_PRECO_%", "VARIACAO_PRECO", "VARIACAO_PRECO (%)", "Variacao Preco (%)"]) or 15  # O
    col_alerta = _col_index_por_header_mc4(ws, ["ALERTA", "Alerta"]) or 19  # S
    col_advertencia = _col_index_por_header_mc4(ws, ["ADVERTENCIA", "Advertencia", "Advertência"]) or 20  # T
    col_intervencao = _col_index_por_header_mc4(ws, ["INTERVENCAO_ANALISTA", "Intervencao Analista", "Intervenção Analista"]) or 21  # U
    col_link_ajuste = _col_index_por_header_mc4(ws, ["LINK_ABRIR_AJUSTE", "Link Abrir Ajuste"]) or 24  # X

    # Coluna CUSTO_ARBITRADO_70: retirar apenas fundo abaixo do cabecalho.
    if col_custo:
        for r in range(2, true_last_row + 1):
            _sem_preenchimento(ws.cell(r, col_custo))
        alteracoes.append("CUSTO_ARBITRADO_70: fundo removido abaixo do cabecalho")
    else:
        alteracoes.append("CUSTO_ARBITRADO_70: coluna nao localizada por cabecalho")

    # Coluna O / VARIACAO_PRECO (%): abaixo 100% ouro; acima 100% vermelho; sem fundo.
    for r in range(2, true_last_row + 1):
        cell = ws.cell(r, col_variacao)
        _sem_preenchimento(cell)
        pct = _parse_percentual_mc4(cell.value)
        if pct is not None:
            if pct < 100:
                cell.font = _font_with_color(cell.font, GOLD)
            elif pct > 100:
                cell.font = _font_with_color(cell.font, RED)
            else:
                cell.font = _font_with_color(cell.font, BLACK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    alteracoes.append("O/VARIACAO_PRECO (%): fundo removido; <100% ouro; >100% vermelho")

    # Colunas S e T: retirar apenas fundo abaixo do cabecalho.
    for col, nome_col in ((col_alerta, "S/ALERTA"), (col_advertencia, "T/ADVERTENCIA")):
        for r in range(2, true_last_row + 1):
            _sem_preenchimento(ws.cell(r, col))
        alteracoes.append(f"{nome_col}: fundo removido abaixo do cabecalho")

    # Coluna U / INTERVENCAO_ANALISTA: maiusculo; NAO verde; SIM vermelho; centralizar.
    for r in range(2, true_last_row + 1):
        cell = ws.cell(r, col_intervencao)
        val = "" if cell.value is None else str(cell.value).strip().upper()
        if val:
            cell.value = val
        if val == "NAO":
            cell.font = _font_with_color(cell.font, GREEN)
        elif val == "SIM":
            cell.font = _font_with_color(cell.font, RED)
        else:
            cell.font = _font_with_color(cell.font, BLACK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    alteracoes.append("U/INTERVENCAO_ANALISTA: maiusculo; NAO verde; SIM vermelho; centralizado")

    # Coluna X / LINK_ABRIR_AJUSTE: ajustar largura.
    try:
        ws.column_dimensions[get_column_letter(col_link_ajuste)].width = 28
        alteracoes.append("X/LINK_ABRIR_AJUSTE: largura ajustada para 28")
    except Exception:
        alteracoes.append("X/LINK_ABRIR_AJUSTE: nao foi possivel ajustar largura")

    return {
        "aplicado": True,
        "linhas_processadas": true_last_row - 1,
        "colunas": {
            "custo_arbitrado_70": col_custo,
            "variacao_preco": col_variacao,
            "alerta": col_alerta,
            "advertencia": col_advertencia,
            "intervencao_analista": col_intervencao,
            "link_abrir_ajuste": col_link_ajuste,
        },
        "alteracoes": alteracoes,
    }




def _extrair_cnpj_emitente_da_chave_nfe_mc4(valor: Any) -> str:
    """Extrai CNPJ do emitente pela chave NF-e: cUF(2)+AAMM(4)+CNPJ(14)."""
    if valor is None:
        return ""
    digits = re.sub(r"\D", "", str(valor))
    if len(digits) >= 20:
        cnpj = digits[6:20]
        if len(cnpj) == 14:
            return cnpj
    return ""


def _aplicar_ajustes_custo_cmv_executivo_auditoria(ws, cfg: Dict[str, Any], mode: Dict[str, Any]) -> Dict[str, Any]:
    """Ajustes visuais da aba CUSTO_CMV.

    A aba concentra CMV e custo derivado do inventario. A rotina nao altera valores,
    apenas melhora leitura: centraliza status, remove fundos excessivos, destaca CMV
    e mantém trilha de cálculo legível.
    """
    nivel = str(cfg.get("layout_nivel") or "").strip().lower()
    modo_calculado = str(mode.get("nivel") or "").strip().lower()
    aplicar = nivel in {"executivo", "auditoria"} or modo_calculado.startswith("executivo") or modo_calculado == "auditoria"
    if ws.title.upper().strip() != "CUSTO_CMV" or not aplicar:
        return {"aplicado": False}

    true_max_col = _true_max_col_by_header(ws)
    true_last_row = _true_last_row_by_values(ws, true_max_col)
    if true_last_row < 2:
        return {"aplicado": True, "linhas": true_last_row, "observacao": "sem linhas de dados"}

    RED = "C00000"; GREEN = "548235"; BLACK = "000000"
    cols = {
        "cmv": _col_index_por_header_mc4(ws, ["CMV"]),
        "situacao": _col_index_por_header_mc4(ws, ["SITUACAO_PRECO", "Situação Preço"]),
        "intervencao": _col_index_por_header_mc4(ws, ["INTERVENCAO_ANALISTA", "Intervenção Analista"]),
        "hash": _col_index_por_header_mc4(ws, ["HASH_LINHA_CALCULO", "Hash Linha Cálculo"]),
        "valores": _col_index_por_header_mc4(ws, ["VALORES_INTERMEDIARIOS", "Valores Intermediários"]),
    }
    alteracoes=[]
    if cols["cmv"]:
        for r in range(2, true_last_row + 1):
            cell = ws.cell(r, cols["cmv"])
            _sem_preenchimento(cell)
            cell.number_format = '#,##0.00'
            cell.font = _font_with_color(cell.font, BLACK, bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
        alteracoes.append("CMV: formato monetario, sem fundo, negrito controlado")
    if cols["situacao"]:
        for r in range(2, true_last_row + 1):
            cell = ws.cell(r, cols["situacao"])
            val = str(cell.value or "").strip().upper()
            if val == "APROVADA":
                cell.font = _font_with_color(cell.font, GREEN, bold=True)
            elif val in {"REVISAR", "BLOQUEADA"}:
                cell.font = _font_with_color(cell.font, RED, bold=True)
            else:
                cell.font = _font_with_color(cell.font, BLACK)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        alteracoes.append("Situacao Preco: centralizada e colorida por criticidade")
    if cols["intervencao"]:
        for r in range(2, true_last_row + 1):
            cell = ws.cell(r, cols["intervencao"])
            val = str(cell.value or "").strip().upper()
            cell.value = val if val else cell.value
            cell.font = _font_with_color(cell.font, RED if val == "SIM" else GREEN if val == "NAO" else BLACK, bold=val in {"SIM", "NAO"})
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        alteracoes.append("Intervencao Analista: padronizada")
    for key in ("hash", "valores"):
        c = cols.get(key)
        if c:
            for r in range(2, true_last_row + 1):
                ws.cell(r, c).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, shrink_to_fit=False)
    return {"aplicado": True, "linhas_processadas": true_last_row - 1, "alteracoes": alteracoes}

def _aplicar_ajustes_pendencias_executivo_auditoria(ws, cfg: Dict[str, Any], mode: Dict[str, Any]) -> Dict[str, Any]:
    """Ajustes microcirurgicos da aba PENDENCIAS.

    Escopo:
    - Somente 04_PENDENCIAS ou PENDENCIAS.
    - Somente layout executivo/auditoria.
    """
    nivel = str(cfg.get("layout_nivel") or "").strip().lower()
    modo_calculado = str(mode.get("nivel") or "").strip().lower()
    aplicar = nivel in {"executivo", "auditoria"} or modo_calculado.startswith("executivo") or modo_calculado == "auditoria"
    nome = ws.title.upper().strip()
    if nome not in {"04_PENDENCIAS", "PENDENCIAS"} or not aplicar:
        return {"aplicado": False}

    true_max_col = _true_max_col_by_header(ws)
    true_last_row = _true_last_row_by_values(ws, true_max_col)
    if true_last_row < 2:
        return {"aplicado": True, "linhas": true_last_row, "observacao": "sem linhas de dados"}

    RED = "C00000"
    alteracoes = []

    # Coluna H / NIVEL: retirar apenas fundo abaixo do cabecalho.
    for r in range(2, true_last_row + 1):
        _sem_preenchimento(ws[f"H{r}"])
    alteracoes.append("H/NIVEL: fundo removido abaixo do cabecalho")

    # Coluna I / INTERVENCAO_ANALISTA: centralizar e fonte vermelha.
    for r in range(2, true_last_row + 1):
        cell = ws[f"I{r}"]
        cell.font = _font_with_color(cell.font, RED)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    alteracoes.append("I/INTERVENCAO_ANALISTA: centralizado e fonte vermelha")

    # Coluna R / LINK_ABRIR_AJUSTE: ajustar largura.
    try:
        ws.column_dimensions["R"].width = 28
        alteracoes.append("R/LINK_ABRIR_AJUSTE: largura ajustada para 28")
    except Exception:
        alteracoes.append("R/LINK_ABRIR_AJUSTE: nao foi possivel ajustar largura")

    return {
        "aplicado": True,
        "linhas_processadas": true_last_row - 1,
        "alteracoes": alteracoes,
    }


def _aplicar_ajustes_fontes_processadas_executivo_auditoria(ws, cfg: Dict[str, Any], mode: Dict[str, Any]) -> Dict[str, Any]:
    """Ajustes visuais enxutos da aba 05_FONTES_PROCESSADAS.

    Não insere colunas e não deriva CNPJ no pós-layout. Lineage deve vir do motor,
    para evitar inchaço e alteração estrutural após a validação técnica.
    """
    nome = ws.title.upper().strip()
    if nome not in {"05_FONTES_PROCESSADAS", "FONTES_PROCESSADAS"}:
        return {"aplicado": False}
    true_max_col = _true_max_col_by_header(ws)
    true_last_row = _true_last_row_by_values(ws, true_max_col)
    for c in range(1, true_max_col + 1):
        h = _norm_header(ws.cell(1, c).value)
        if h in {"HASH_FONTE", "CODIGO_CONTROLE_ARQUIVO"}:
            ws.column_dimensions[get_column_letter(c)].width = 22
            for r in range(2, true_last_row + 1):
                ws.cell(r, c).font = Font(name=cfg["fonte_corpo"], size=8, color=PALETA["muted"])
                ws.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", shrink_to_fit=True)
    return {"aplicado": True, "linhas_processadas": max(true_last_row - 1, 0), "alteracoes": ["FONTES_PROCESSADAS: lineage enxuto, sem colunas derivadas no pós-layout"]}


def _style_sheet(ws, perfil: str, cfg: Dict[str, Any]) -> Dict[str, Any]:
    t0 = time.perf_counter()
    _clear_column_dimensions(ws)

    true_max_col = _true_max_col_by_header(ws)
    true_last_row = _true_last_row_by_values(ws, true_max_col)
    mode = _mode_settings(cfg, true_last_row)

    technical_hdr = _technical_headers(ws, true_max_col)
    header_mapping = _apply_friendly_headers(ws, true_max_col, cfg)
    display_hdr = [str(ws.cell(1, c).value or "").strip() for c in range(1, true_max_col + 1)]

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TAB_COLORS.get(ws.title, PALETA["slate"])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(true_max_col)}{true_last_row}"
    ws.sheet_view.zoomScale = 90

    header_fill = PatternFill("solid", fgColor=HEADER_COLORS.get(ws.title, PALETA["slate"]))
    header_border = Border(bottom=Side(style="medium", color=PALETA["white"]))
    bottom_line = Border(bottom=Side(style="hair", color=PALETA["line"]))
    zebra_fill = PatternFill("solid", fgColor=PALETA["zebra"])
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    status_cols = [i + 1 for i, h in enumerate(technical_hdr) if any(x in h for x in ("STATUS", "SITUACAO", "NIVEL", "RESULTADO"))]
    chave_cols = [i + 1 for i, h in enumerate(technical_hdr) if "CHAVE" in h]
    hash_cols = [i + 1 for i, h in enumerate(technical_hdr) if "SHA256" in h or "HASH" in h]
    detail_cols = [i + 1 for i, h in enumerate(technical_hdr) if any(x in h for x in ("DETALHE", "MOTIVO", "JUSTIFICATIVA", "OBS", "TRAIL", "VALORES_INTERMEDIARIOS", "ACAO"))]

    ws.row_dimensions[1].height = _int(cfg.get("altura_linha_cabecalho"), 24)
    for c in range(1, true_max_col + 1):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = Font(name=cfg["fonte_cabecalho"], size=cfg["tamanho_cabecalho"], bold=True, color=PALETA["white"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    rows_to_style = int(mode["body_style_rows"])
    body_height = _int(cfg.get("altura_linha_corpo"), 18)
    for r in range(2, rows_to_style + 1):
        ws.row_dimensions[r].height = body_height
        fill = zebra_fill if (mode["zebra"] and r % 2 == 0) else white_fill
        for c in range(1, true_max_col + 1):
            cell = ws.cell(r, c)
            # In modo rápido, RAW já saiu com Aptos Narrow 10; aqui só corrigimos as primeiras linhas visíveis.
            cell.fill = fill
            if mode["deep"]:
                cell.border = bottom_line
            cell.font = Font(name=cfg["fonte_corpo"], size=cfg["tamanho_corpo"], color=PALETA["text"])
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for c in detail_cols:
        for r in range(2, rows_to_style + 1):
            ws.cell(r, c).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for c in chave_cols:
        for r in range(2, rows_to_style + 1):
            cell = ws.cell(r, c)
            cell.number_format = "@"
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for c in hash_cols:
        for r in range(2, rows_to_style + 1):
            cell = ws.cell(r, c)
            cell.number_format = "@"
            cell.font = Font(name="Aptos Narrow", size=8, color=PALETA["muted"])
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=True)

    if mode["status"]:
        for c in status_cols:
            for r in range(2, min(true_last_row, rows_to_style) + 1):
                fill = _status_fill(ws.cell(r, c).value)
                if fill:
                    ws.cell(r, c).fill = fill
                    ws.cell(r, c).font = Font(name=cfg["fonte_corpo"], size=cfg["tamanho_corpo"], bold=True, color=PALETA["text"])

    sample_rows = min(true_last_row, 300 if not mode["deep"] else 1000)
    max_width = _int(cfg.get("max_largura_coluna"), 72)
    for c in range(1, true_max_col + 1):
        display_header = display_hdr[c - 1]
        tech_header = technical_hdr[c - 1]
        max_len = len(display_header)
        for r in range(2, sample_rows + 1):
            val = ws.cell(r, c).value
            if val is None:
                continue
            max_len = max(max_len, max((len(line) for line in str(val).splitlines()), default=0))
        ws.column_dimensions[get_column_letter(c)].width = _column_width(display_header, tech_header, max_len, max_width)

    _clean_residual_styles(ws, true_last_row, true_max_col, cfg, bool(mode["deep"]))

    capa_ajustes = _aplicar_ajustes_capa_executivo_auditoria(ws, cfg, mode)
    movimento_ajustes = _aplicar_ajustes_movimento_validado_executivo_auditoria(ws, cfg, mode)
    arbitramento_ajustes = _aplicar_ajustes_arbitramento_executivo_auditoria(ws, cfg, mode)
    custo_cmv_ajustes = _aplicar_ajustes_custo_cmv_executivo_auditoria(ws, cfg, mode)
    pendencias_ajustes = _aplicar_ajustes_pendencias_executivo_auditoria(ws, cfg, mode)
    fontes_processadas_ajustes = _aplicar_ajustes_fontes_processadas_executivo_auditoria(ws, cfg, mode)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    try:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        pass

    return {
        "aba": ws.title,
        "linhas": true_last_row,
        "colunas": true_max_col,
        "modo": mode["nivel"],
        "linhas_estilizadas": rows_to_style,
        "segundos": round(time.perf_counter() - t0, 3),
        "cabecalhos_amigaveis": header_mapping,
        "ajustes_capa": capa_ajustes,
        "ajustes_movimento_validado": movimento_ajustes,
        "ajustes_arbitramento": arbitramento_ajustes,
        "ajustes_custo_cmv": custo_cmv_ajustes,
        "ajustes_pendencias": pendencias_ajustes,
        "ajustes_fontes_processadas": fontes_processadas_ajustes,
    }


def aplicar_layout_premium(
    input_path: str | Path,
    output_path: str | Path | None = None,
    perfil: str = "output_fiscal",
    config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    cfg = dict(DEFAULT_CONFIG)
    if config:
        cfg.update(config)

    input_path = Path(input_path)
    output_path = Path(output_path) if output_path else input_path

    if not input_path.exists():
        raise FileNotFoundError(f"Arquivo de entrada nao localizado: {input_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    start_all = time.perf_counter()
    before_hash = sha256_file(input_path)

    if input_path.resolve() != output_path.resolve():
        shutil.copy2(input_path, output_path)

    wb = load_workbook(output_path)
    try:
        ordenar_pre = _as_bool(cfg.get("ordenar_chave_nfe_pre_validacao"), True)
        ordenar_mov = _as_bool(cfg.get("ordenar_chave_nfe_movimento"), False)

        abas_info = []
        for ws in wb.worksheets:
            true_max_col = _true_max_col_by_header(ws)
            if perfil == "pre_validacao" and ordenar_pre and ws.title.upper() == "FONTES_ANALISADAS":
                _sort_by_chave_nfe(ws, true_max_col)
            if perfil == "output_fiscal" and ordenar_mov and ws.title.upper() == "01_MOVIMENTO_VALIDADO":
                _sort_by_chave_nfe(ws, true_max_col)
            abas_info.append(_style_sheet(ws, perfil, cfg))

        wb.properties.creator = wb.properties.creator or "MC4 CONTABILIDADE"
        wb.properties.lastModifiedBy = "MC4 CONTABILIDADE"
        wb.properties.title = wb.properties.title or "Motor Fiscal RIR70 by MC4"
        wb.save(output_path)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    after_hash = sha256_file(output_path)
    result = {
        "layout_mc4": "OK",
        "versao_layout": "1.3.28-producao-forense",
        "perfil": perfil,
        "layout_nivel": str(cfg.get("layout_nivel") or "rapido"),
        "arquivo_entrada": str(input_path),
        "arquivo_saida": str(output_path),
        "sha256_antes_layout": before_hash,
        "sha256_depois_layout": after_hash,
        "aplicado_em": datetime.now().isoformat(timespec="seconds"),
        "segundos_total": round(time.perf_counter() - start_all, 3),
        "abas": abas_info,
        "observacao": "Layout performance aplicado. Ajustes especiais na CAPA, MOVIMENTO_VALIDADO, ARBITRAMENTO, CUSTO_CMV, PENDENCIAS e FONTES_PROCESSADAS para modo executivo/auditoria.",
    }
    sidecar = output_path.with_suffix(output_path.suffix + ".layout_mc4.json")
    sidecar.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Aplica layout premium MC4 v1.3.28-producao-forense em arquivo Excel.")
    parser.add_argument("input_path")
    parser.add_argument("output_path", nargs="?")
    parser.add_argument("--perfil", default="output_fiscal")
    parser.add_argument("--nivel", default=None)
    args = parser.parse_args()
    cfg = {}
    if args.nivel:
        cfg["layout_nivel"] = args.nivel
    print(json.dumps(aplicar_layout_premium(args.input_path, args.output_path, perfil=args.perfil, config=cfg), ensure_ascii=False, indent=2))
