# -*- coding: utf-8 -*-
"""
sistema_rir70_web.py - Interface web analitica e performatica para o Motor Fiscal RIR70 by MC4.
Versao: 1.3.28-producao-forense | MC4 CONTABILIDADE E GESTAO DE NEGOCIOS
CNPJ 09.944.432/0001-25

Camada web local com navegacao por abas, cache SQLite, paginacao server-side,
filtros no backend, drill-through e painel executivo.
Ferramenta proprietaria MC4. Uso interno licenciado.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import sqlite3
import subprocess
import sys
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from flask import Flask, Response, jsonify, redirect, render_template_string, request, send_file, url_for
except ImportError:
    print("Flask nao instalado. Execute: pip install flask")
    sys.exit(1)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

APP_VERSION = "1.3.28-producao-forense"
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"
INPUT_DIR = BASE_DIR / "input"
SCRIPT = BASE_DIR / "motor_arbitramento.py"
CONFIG_PATH = BASE_DIR / "config_arbitramento_rir70.json"
CACHE_DB = OUTPUT_DIR / ".mc4_rir70_web_cache.sqlite"

MOTOR_OWNER = "MC4 CONTABILIDADE E GESTAO DE NEGOCIOS"
MOTOR_OWNER_CNPJ = "09.944.432/0001-25"
MOTOR_BRAND = "Motor Fiscal RIR70 by MC4"
MOTOR_LICENSED_TO = "101 Brasil Industria de Bebidas Ltda"

SHEET_MAP = {
    "capa": "CAPA",
    "movimento": "01_MOVIMENTO_VALIDADO",
    "arbitramento": "02_ARBITRAMENTO",
    "inventario": "03_INVENTARIO_VALORIZADO",
    "pendencias": "04_PENDENCIAS",
    "fontes": "05_FONTES_PROCESSADAS",
}
SHEET_LABELS = {
    "capa": "CAPA",
    "movimento": "Movimento Validado",
    "arbitramento": "Arbitramento",
    "inventario": "Inventario Valorizado",
    "pendencias": "Pendencias",
    "fontes": "Fontes Processadas",
}
SHEET_HELP = {
    "capa": "Resumo executivo, status geral, periodo, licenca MC4 e proximos passos.",
    "movimento": "Base fiscal itemizada lida, saneada e classificada pelo motor.",
    "arbitramento": "Memoria mensal do maior preco de venda e custo arbitrado RIR70.",
    "inventario": "Inventario valorizado pelo criterio de arbitramento.",
    "pendencias": "Fila operacional de revisao, correcao e aprovacao fiscal.",
    "fontes": "Trilha das fontes processadas, arquivos, sheets e contagens.",
}
PREFERRED_COLUMNS = {
    "pendencias": ["EMPRESA", "MES", "CODIGO_PRODUTO", "DESCRICAO_PRODUTO", "TIPO_PENDENCIA", "NIVEL", "ACAO_ANALISTA", "SITUACAO", "ALTERA_CALCULO"],
    "arbitramento": ["EMPRESA", "MES", "CODIGO_PRODUTO", "DESCRICAO_PRODUTO", "NCM", "MAIOR_PRECO_UNITARIO", "CUSTO_ARBITRADO_70", "QTD_VENDAS", "ORIGEM_DO_PRECO"],
    "inventario": ["EMPRESA", "MES", "CODIGO_PRODUTO", "DESCRICAO_PRODUTO", "NCM", "QTDE_INVENTARIO", "CUSTO_ARBITRADO_70", "VALOR_ESTOQUE_ARBITRADO", "INTERVENCAO_ANALISTA", "ACAO_ANALISTA"],
    "movimento": ["EMPRESA", "MES", "CHAVE_NFE", "NUMERO_NF", "CODIGO_PRODUTO", "DESCRICAO_PRODUTO", "NCM", "CFOP", "QTD_CALCULO", "VALOR_PRODUTO_BASE", "PARTICIPA_CALCULO", "MOTIVO_CLASSIFICACAO"],
    "fontes": ["ARQUIVO", "SHEET", "TIPO_FONTE", "REGISTROS_LIDOS", "REGISTROS_VALIDOS", "REGISTROS_DUPLICADOS", "REGISTROS_BLOQUEADOS_CFOP", "STATUS"],
}
FILTER_ALIASES = {
    "empresa": ["EMPRESA", "CNPJ", "CNPJ_EMITENTE", "CNPJ Emitente", "CNPJDOEMITENTE"],
    "mes": ["MES", "COMPETENCIA", "PERIODO", "COMP"],
    "nivel": ["NIVEL", "STATUS", "SITUACAO", "SITUAÇÃO", "SITUACAO_ANALISTA"],
    "tipo": ["TIPO_PENDENCIA", "TIPO", "MOTIVO", "MOTIVO_CLASSIFICACAO"],
}

app = Flask(__name__)
OUTPUT_DIR.mkdir(exist_ok=True)

_run_lock = threading.Lock()
_run_status: Dict[str, Any] = {"running": False, "exitcode": None, "started_at": None, "finished_at": None, "message": "Aguardando execucao"}
_run_env_vars: Dict[str, str] = {}
_run_output_tail: List[str] = []
_last_form: Dict[str, str] = {}

_cache_lock = threading.Lock()
_cache_job: Dict[str, Any] = {"running": False, "status": "idle", "message": "Cache nao iniciado", "started_at": None, "finished_at": None, "error": None, "rows": 0, "sheet": None}


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    if text.lower() in {"none", "nan", "nat"}:
        return ""
    return text


def norm(value: Any) -> str:
    text = safe_text(value).upper()
    text = re.sub(r"[^A-Z0-9]+", "", text)
    return text


def format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float):
        # Preserva chave/NF e numeros em texto quando vierem do Excel como string; floats aqui sao valores calculados.
        if abs(value) >= 1000:
            return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"{value:.4f}".rstrip("0").rstrip(".").replace(".", ",")
    if isinstance(value, int):
        return str(value)
    return safe_text(value)


def last_output_path() -> Optional[Path]:
    p = OUTPUT_DIR / "ultimo_arquivo_gerado.txt"
    if p.exists():
        try:
            xlsx = Path(p.read_text(encoding="utf-8", errors="replace").strip())
            if xlsx.exists():
                return xlsx
        except Exception:
            pass
    if OUTPUT_DIR.exists():
        files = sorted(OUTPUT_DIR.glob("Arbitramento_*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True)
        if files:
            return files[0]
    return None


def xlsx_fingerprint(path: Optional[Path]) -> Dict[str, Any]:
    if not path or not path.exists():
        return {"path": "", "mtime": 0, "size": 0, "name": "", "sha256": ""}
    st = path.stat()
    return {"path": str(path.resolve()), "mtime": int(st.st_mtime), "size": st.st_size, "name": path.name, "sha256": file_sha256(path)}


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    try:
        with path.open("rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return ""


def load_json(path: Path) -> Dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8", errors="replace"))
    except Exception:
        return {}


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(CACHE_DB, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA temp_store=MEMORY")
    conn.execute("PRAGMA cache_size=-50000")
    return conn


def init_cache_schema(conn: sqlite3.Connection) -> None:
    conn.execute("CREATE TABLE IF NOT EXISTS cache_state (id INTEGER PRIMARY KEY CHECK (id=1), xlsx_path TEXT, xlsx_mtime INTEGER, xlsx_size INTEGER, xlsx_sha256 TEXT, cache_sha256 TEXT, built_at TEXT, status TEXT, app_version TEXT, aviso TEXT)")
    conn.execute("CREATE TABLE IF NOT EXISTS sheet_meta (sheet_key TEXT PRIMARY KEY, sheet_name TEXT, label TEXT, headers_json TEXT, row_count INTEGER, filter_map_json TEXT)")
    conn.execute("CREATE TABLE IF NOT EXISTS sheet_summary (sheet_key TEXT, kind TEXT, label TEXT, value INTEGER, PRIMARY KEY(sheet_key, kind, label))")
    conn.commit()


def cache_current() -> bool:
    xlsx = last_output_path()
    fp = xlsx_fingerprint(xlsx)
    if not fp["path"] or not CACHE_DB.exists():
        return False
    try:
        with get_conn() as conn:
            init_cache_schema(conn)
            row = conn.execute("SELECT * FROM cache_state WHERE id=1").fetchone()
            if not row:
                return False
            return row["xlsx_path"] == fp["path"] and int(row["xlsx_mtime"] or 0) == fp["mtime"] and int(row["xlsx_size"] or 0) == fp["size"] and safe_text(row["xlsx_sha256"]) == fp.get("sha256", "") and row["status"] == "OK"
    except Exception:
        return False


def cache_status_payload() -> Dict[str, Any]:
    xlsx = last_output_path()
    fp = xlsx_fingerprint(xlsx)
    current = cache_current()
    state = {}
    if CACHE_DB.exists():
        try:
            with get_conn() as conn:
                init_cache_schema(conn)
                row = conn.execute("SELECT * FROM cache_state WHERE id=1").fetchone()
                if row:
                    state = dict(row)
        except Exception as exc:
            state = {"status": "ERRO", "erro": str(exc)}
    return {"xlsx": fp, "cache_db": str(CACHE_DB), "current": current, "state": state, "job": dict(_cache_job)}


def unique_headers(values: Iterable[Any]) -> List[str]:
    used: Dict[str, int] = {}
    headers: List[str] = []
    for idx, raw in enumerate(values):
        name = safe_text(raw) or f"COLUNA_{idx+1}"
        if name in used:
            used[name] += 1
            name = f"{name}_{used[name]}"
        else:
            used[name] = 1
        headers.append(name)
    return headers


def filter_map_for_headers(headers: List[str]) -> Dict[str, Optional[str]]:
    nmap = {norm(h): h for h in headers}
    out: Dict[str, Optional[str]] = {}
    for key, aliases in FILTER_ALIASES.items():
        found = None
        for alias in aliases:
            if norm(alias) in nmap:
                found = nmap[norm(alias)]
                break
        out[key] = found
    return out


def col_name(idx: int) -> str:
    return f"c{idx+1:03d}"


def build_search_text(values: List[str]) -> str:
    return " ".join(v.upper() for v in values if v)


def drop_sheet_table(conn: sqlite3.Connection, sheet_key: str) -> None:
    conn.execute(f"DROP TABLE IF EXISTS sheet_{sheet_key}")


def create_sheet_table(conn: sqlite3.Connection, sheet_key: str, ncols: int) -> None:
    cols = ", ".join(f"{col_name(i)} TEXT" for i in range(ncols))
    conn.execute(f"CREATE TABLE sheet_{sheet_key} (rid INTEGER PRIMARY KEY, search_text TEXT, {cols})")
    conn.execute(f"CREATE INDEX idx_{sheet_key}_search ON sheet_{sheet_key}(search_text)")


def insert_sheet_rows(conn: sqlite3.Connection, sheet_key: str, headers: List[str], rows_iter: Iterable[Tuple[int, Tuple[Any, ...]]], filter_map: Dict[str, Optional[str]]) -> int:
    ncols = len(headers)
    fields = ["rid", "search_text"] + [col_name(i) for i in range(ncols)]
    placeholders = ",".join("?" for _ in fields)
    sql = f"INSERT INTO sheet_{sheet_key} ({','.join(fields)}) VALUES ({placeholders})"
    batch: List[Tuple[Any, ...]] = []
    count = 0
    for rid, row in rows_iter:
        values = [format_value(row[i]) if i < len(row) else "" for i in range(ncols)]
        batch.append(tuple([rid, build_search_text(values)] + values))
        count += 1
        if len(batch) >= 1500:
            conn.executemany(sql, batch)
            conn.commit()
            batch.clear()
            with _cache_lock:
                _cache_job["rows"] = _cache_job.get("rows", 0) + 1500
    if batch:
        conn.executemany(sql, batch)
        conn.commit()
        with _cache_lock:
            _cache_job["rows"] = _cache_job.get("rows", 0) + len(batch)
    # indices de filtros comuns quando as colunas existem
    h_to_idx = {h: i for i, h in enumerate(headers)}
    for fname, header in filter_map.items():
        if header and header in h_to_idx:
            conn.execute(f"CREATE INDEX IF NOT EXISTS idx_{sheet_key}_{fname} ON sheet_{sheet_key}({col_name(h_to_idx[header])})")
    conn.commit()
    return count


def compute_summary_from_table(conn: sqlite3.Connection, sheet_key: str, headers: List[str], filter_map: Dict[str, Optional[str]]) -> None:
    conn.execute("DELETE FROM sheet_summary WHERE sheet_key=?", (sheet_key,))
    h_to_idx = {h: i for i, h in enumerate(headers)}
    def field(header: Optional[str]) -> Optional[str]:
        return col_name(h_to_idx[header]) if header and header in h_to_idx else None
    for kind in ["empresa", "mes", "nivel", "tipo"]:
        f = field(filter_map.get(kind))
        if not f:
            continue
        # Limita distribuicoes grandes, mantendo top 200.
        rows = conn.execute(f"SELECT {f} AS label, COUNT(*) AS value FROM sheet_{sheet_key} WHERE TRIM(COALESCE({f},''))<>'' GROUP BY {f} ORDER BY value DESC LIMIT 200").fetchall()
        conn.executemany("INSERT OR REPLACE INTO sheet_summary(sheet_key, kind, label, value) VALUES(?,?,?,?)", [(sheet_key, kind, r["label"], int(r["value"])) for r in rows])
    conn.commit()


def cache_db_sha256() -> str:
    try:
        if not CACHE_DB.exists():
            return ""
        return file_sha256(CACHE_DB)
    except Exception:
        return ""

def build_cache_sync() -> None:
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl nao instalado")
    xlsx = last_output_path()
    if not xlsx or not xlsx.exists():
        raise RuntimeError("Nenhum Excel de output localizado em output/")
    fp = xlsx_fingerprint(xlsx)
    with _cache_lock:
        _cache_job.update({"running": True, "status": "running", "message": "Abrindo Excel para indexacao", "started_at": datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "finished_at": None, "error": None, "rows": 0, "sheet": None})
    try:
        if CACHE_DB.exists():
            CACHE_DB.unlink()
        conn = get_conn()
        init_cache_schema(conn)
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        try:
            for sheet_key, sheet_name in SHEET_MAP.items():
                with _cache_lock:
                    _cache_job.update({"sheet": sheet_name, "message": f"Indexando {sheet_name}"})
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
                headers = unique_headers(header_row)
                if not headers:
                    headers = ["COLUNA_1"]
                filter_map = filter_map_for_headers(headers)
                drop_sheet_table(conn, sheet_key)
                create_sheet_table(conn, sheet_key, len(headers))
                count = insert_sheet_rows(conn, sheet_key, headers, ((rn, row) for rn, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2)), filter_map)
                compute_summary_from_table(conn, sheet_key, headers, filter_map)
                conn.execute("INSERT OR REPLACE INTO sheet_meta(sheet_key, sheet_name, label, headers_json, row_count, filter_map_json) VALUES(?,?,?,?,?,?)", (sheet_key, sheet_name, SHEET_LABELS.get(sheet_key, sheet_name), json.dumps(headers, ensure_ascii=False), count, json.dumps(filter_map, ensure_ascii=False)))
                conn.commit()
        finally:
            try:
                wb.close()
            except Exception:
                pass
        aviso = "CACHE_ANALITICO_DERIVADO: o Excel oficial continua sendo a evidencia fiscal. Reindexe sempre que o Excel mudar."
        conn.execute("INSERT OR REPLACE INTO cache_state(id, xlsx_path, xlsx_mtime, xlsx_size, xlsx_sha256, cache_sha256, built_at, status, app_version, aviso) VALUES(1,?,?,?,?,?,?,?,?,?)", (fp["path"], fp["mtime"], fp["size"], fp.get("sha256", ""), "PENDENTE_FECHAMENTO", datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "OK", APP_VERSION, aviso))
        conn.commit()
        digest = cache_db_sha256()
        conn.execute("UPDATE cache_state SET cache_sha256=? WHERE id=1", (digest,))
        conn.commit()
        conn.close()
        with _cache_lock:
            _cache_job.update({"running": False, "status": "OK", "message": "Cache concluido", "finished_at": datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "error": None, "sheet": None})
    except Exception as exc:
        with _cache_lock:
            _cache_job.update({"running": False, "status": "ERRO", "message": "Falha ao indexar Excel", "finished_at": datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "error": str(exc)})
        raise


def start_cache_build(force: bool = False) -> Dict[str, Any]:
    with _cache_lock:
        if _cache_job.get("running"):
            return dict(_cache_job)
    if cache_current() and not force:
        return cache_status_payload()["job"]
    def worker():
        try:
            build_cache_sync()
        except Exception:
            pass
    t = threading.Thread(target=worker, daemon=True)
    t.start()
    return dict(_cache_job)


def get_meta() -> Dict[str, Any]:
    if not cache_current():
        return {"sheets": {}, "cache_current": False, "xlsx": xlsx_fingerprint(last_output_path())}
    with get_conn() as conn:
        init_cache_schema(conn)
        rows = conn.execute("SELECT * FROM sheet_meta ORDER BY sheet_key").fetchall()
        sheets = {}
        for r in rows:
            sheets[r["sheet_key"]] = {
                "key": r["sheet_key"],
                "name": r["sheet_name"],
                "label": r["label"],
                "headers": json.loads(r["headers_json"] or "[]"),
                "row_count": int(r["row_count"] or 0),
                "filter_map": json.loads(r["filter_map_json"] or "{}"),
            }
        state = conn.execute("SELECT * FROM cache_state WHERE id=1").fetchone()
        return {"sheets": sheets, "cache_current": True, "xlsx": xlsx_fingerprint(last_output_path()), "state": dict(state) if state else {}}


def load_summary(sheet_key: str = "pendencias") -> Dict[str, Any]:
    out = {"total": 0, "critico": 0, "revisar": 0, "ok": 0, "por_tipo": [], "por_empresa": [], "por_mes": [], "por_nivel": []}
    if not cache_current():
        return out
    with get_conn() as conn:
        meta = conn.execute("SELECT row_count FROM sheet_meta WHERE sheet_key=?", (sheet_key,)).fetchone()
        out["total"] = int(meta["row_count"] or 0) if meta else 0
        nivel_rows = conn.execute("SELECT label, value FROM sheet_summary WHERE sheet_key=? AND kind='nivel' ORDER BY value DESC", (sheet_key,)).fetchall()
        for r in nivel_rows:
            lab = (r["label"] or "").upper()
            val = int(r["value"] or 0)
            if "CRITICO" in lab:
                out["critico"] += val
            elif "REVISAR" in lab:
                out["revisar"] += val
            else:
                out["ok"] += val
        for kind, out_key in [("tipo", "por_tipo"), ("empresa", "por_empresa"), ("mes", "por_mes"), ("nivel", "por_nivel")]:
            rows = conn.execute("SELECT label, value FROM sheet_summary WHERE sheet_key=? AND kind=? ORDER BY value DESC LIMIT 12", (sheet_key, kind)).fetchall()
            out[out_key] = [{"label": r["label"], "value": int(r["value"] or 0)} for r in rows]
    return out


def sheet_query(sheet_key: str, page: int, per_page: int, search: str, filters: Dict[str, str], mode: str) -> Dict[str, Any]:
    if not cache_current():
        return {"cache_current": False, "headers": [], "rows": [], "total": 0, "page": page, "per_page": per_page, "pages": 0, "distinct": {}}
    with get_conn() as conn:
        meta_row = conn.execute("SELECT * FROM sheet_meta WHERE sheet_key=?", (sheet_key,)).fetchone()
        if not meta_row:
            return {"cache_current": True, "headers": [], "rows": [], "total": 0, "page": page, "per_page": per_page, "pages": 0, "distinct": {}}
        headers = json.loads(meta_row["headers_json"] or "[]")
        filter_map = json.loads(meta_row["filter_map_json"] or "{}")
        h_to_idx = {h: i for i, h in enumerate(headers)}
        selected = headers
        if mode == "executivo" and sheet_key in PREFERRED_COLUMNS:
            selected = [h for h in PREFERRED_COLUMNS[sheet_key] if h in h_to_idx] or headers[:20]
        elif mode == "compacto":
            selected = headers[:15]
        where: List[str] = []
        params: List[Any] = []
        if search.strip():
            where.append("search_text LIKE ?")
            params.append(f"%{search.strip().upper()}%")
        for fkey, fval in filters.items():
            if not fval:
                continue
            header = filter_map.get(fkey)
            if header and header in h_to_idx:
                where.append(f"{col_name(h_to_idx[header])} = ?")
                params.append(fval)
        where_sql = (" WHERE " + " AND ".join(where)) if where else ""
        table = f"sheet_{sheet_key}"
        total = int(conn.execute(f"SELECT COUNT(*) AS n FROM {table}{where_sql}", params).fetchone()["n"])
        page = max(page, 1)
        per_page = max(min(per_page, 500), 25)
        offset = (page - 1) * per_page
        select_fields = ["rid"] + [col_name(h_to_idx[h]) + f" AS '{h}'" for h in selected]
        rows = [dict(r) for r in conn.execute(f"SELECT {', '.join(select_fields)} FROM {table}{where_sql} ORDER BY rid LIMIT ? OFFSET ?", params + [per_page, offset]).fetchall()]
        distinct: Dict[str, List[str]] = {}
        for kind in ["empresa", "mes", "nivel", "tipo"]:
            summary = conn.execute("SELECT label FROM sheet_summary WHERE sheet_key=? AND kind=? ORDER BY value DESC, label LIMIT 300", (sheet_key, kind)).fetchall()
            distinct[kind] = [r["label"] for r in summary]
        return {"cache_current": True, "sheet_key": sheet_key, "label": SHEET_LABELS.get(sheet_key, sheet_key), "help": SHEET_HELP.get(sheet_key, ""), "headers": selected, "all_headers": headers, "rows": rows, "total": total, "page": page, "per_page": per_page, "pages": (total + per_page - 1) // per_page if total else 0, "distinct": distinct, "mode": mode}


def read_row(sheet_key: str, rid: int) -> Dict[str, Any]:
    if not cache_current():
        return {}
    with get_conn() as conn:
        meta = conn.execute("SELECT headers_json FROM sheet_meta WHERE sheet_key=?", (sheet_key,)).fetchone()
        if not meta:
            return {}
        headers = json.loads(meta["headers_json"] or "[]")
        fields = [col_name(i) + f" AS '{h}'" for i, h in enumerate(headers)]
        row = conn.execute(f"SELECT rid, {', '.join(fields)} FROM sheet_{sheet_key} WHERE rid=?", (rid,)).fetchone()
        return dict(row) if row else {}


def integrity_summary() -> Dict[str, Any]:
    manifest = load_json(BASE_DIR / "MANIFESTO_PACOTE_MC4.json")
    rows = []
    divergentes = []
    for item in manifest.get("arquivos", []) if isinstance(manifest, dict) else []:
        rel = item.get("nome") or ""
        expected = item.get("sha256") or ""
        found = file_sha256(BASE_DIR / rel)
        status = "OK" if expected and expected == found else "DIVERGENTE"
        if status != "OK":
            divergentes.append(rel)
        rows.append({"arquivo": rel, "status": status, "esperado": expected, "encontrado": found})
    return {"status": "OK" if rows and not divergentes else "VIOLADA", "arquivos": rows, "divergentes": divergentes, "manifesto": manifest}


def load_log_lines() -> List[str]:
    if not OUTPUT_DIR.exists():
        return _run_output_tail[-200:]
    logs = sorted(OUTPUT_DIR.glob("*.log"), key=lambda p: p.stat().st_mtime, reverse=True)
    if logs:
        try:
            lines = logs[0].read_text(encoding="utf-8", errors="replace").splitlines()
            return lines[-500:]
        except Exception:
            pass
    return _run_output_tail[-200:]


def run_motor_thread() -> None:
    global _run_status, _run_output_tail
    with _run_lock:
        _run_status.update({"running": True, "exitcode": None, "started_at": datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "finished_at": None, "message": "Motor em execucao"})
    try:
        env = {**os.environ, "RIR70_EXECUTADO_PELO_BAT": "0"}
        env.update(_run_env_vars)
        proc = subprocess.Popen([sys.executable, str(SCRIPT)], cwd=str(BASE_DIR), stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding="utf-8", errors="replace", env=env)
        tail: List[str] = []
        assert proc.stdout is not None
        for line in proc.stdout:
            tail.append(line.rstrip("\n"))
            _run_output_tail = tail[-300:]
        code = proc.wait()
        _run_status.update({"running": False, "exitcode": code, "finished_at": datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "message": "Concluido" if code == 0 else "Erro na execucao"})
        # Invalida cache para que a nova saida seja indexada.
        if code == 0 and CACHE_DB.exists():
            try:
                CACHE_DB.unlink()
            except Exception:
                pass
    except Exception as exc:
        _run_output_tail.append(f"Erro ao iniciar motor: {exc}")
        _run_status.update({"running": False, "exitcode": -1, "finished_at": datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "message": str(exc)})


HTML = r"""
<!doctype html>
<html lang="pt-br">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Motor Fiscal RIR70 by MC4</title>
<style>
:root{
  --nav:#132B4F;--nav2:#1E3A67;--accent:#F59E0B;--bg:#F3F6FA;--card:#FFFFFF;--line:#D8E0EA;--text:#101828;--muted:#667085;
  --red:#DC2626;--red-bg:#FEE2E2;--amber:#D97706;--amber-bg:#FEF3C7;--green:#16A34A;--green-bg:#DCFCE7;--blue:#2563EB;--shadow:0 10px 30px rgba(16,24,40,.08)
}
*{box-sizing:border-box}body{margin:0;font-family:'Arial Narrow',Arial,sans-serif;background:var(--bg);color:var(--text);font-size:14px}button,input,select{font-family:inherit}.app{min-height:100vh;display:grid;grid-template-columns:280px 1fr}.side{background:linear-gradient(180deg,var(--nav),#0B1C35);color:white;position:sticky;top:0;height:100vh;overflow:auto}.brand{padding:22px 20px;border-bottom:1px solid rgba(255,255,255,.14)}.brand h1{font-size:20px;line-height:1.15;margin:0 0 10px}.brand p{margin:0;color:#C7D7EA;font-size:12px;line-height:1.45}.nav{padding:12px 0}.nav button{width:100%;border:0;background:transparent;color:#EAF2FF;padding:12px 18px;text-align:left;display:flex;gap:10px;align-items:center;cursor:pointer;border-left:4px solid transparent;font-weight:700}.nav button:hover,.nav button.active{background:rgba(255,255,255,.08);border-left-color:var(--accent)}.nav .section{padding:16px 18px 6px;color:#99B3D1;font-size:11px;text-transform:uppercase;letter-spacing:.08em}.main{min-width:0}.top{height:66px;background:white;border-bottom:1px solid var(--line);display:flex;align-items:center;justify-content:space-between;padding:0 28px;position:sticky;top:0;z-index:10}.top h2{font-size:21px;margin:0}.top-actions{display:flex;gap:8px;align-items:center}.content{padding:24px 28px 56px}.notice{border-left:5px solid var(--accent);background:#FFFBEB;padding:13px 16px;border-radius:12px;margin-bottom:18px;box-shadow:0 1px 2px rgba(16,24,40,.04)}.grid{display:grid;gap:16px}.cards{grid-template-columns:repeat(4,minmax(0,1fr))}.card{background:var(--card);border:1px solid var(--line);border-radius:18px;box-shadow:0 1px 2px rgba(16,24,40,.04);overflow:hidden}.card-h{padding:15px 17px;border-bottom:1px solid var(--line);display:flex;align-items:center;justify-content:space-between;gap:12px;background:#FBFCFE}.card-h h3{font-size:16px;margin:0}.card-b{padding:16px}.metric{font-size:34px;font-weight:900;letter-spacing:-.03em;line-height:1}.metric small{display:block;font-size:12px;font-weight:700;color:var(--muted);letter-spacing:0;margin-top:5px}.crit{color:var(--red)}.rev{color:var(--amber)}.ok{color:var(--green)}.blue{color:var(--blue)}.pill{display:inline-flex;align-items:center;gap:5px;padding:4px 9px;border-radius:999px;font-size:12px;font-weight:900}.pill.crit{background:var(--red-bg);color:#991B1B}.pill.rev{background:var(--amber-bg);color:#92400E}.pill.ok{background:var(--green-bg);color:#166534}.pill.muted{background:#EEF2F7;color:#475467}.btn{height:38px;border:0;border-radius:10px;padding:0 14px;background:var(--nav);color:white;font-weight:900;cursor:pointer;display:inline-flex;align-items:center;justify-content:center;gap:8px;text-decoration:none}.btn:hover{filter:brightness(1.08)}.btn.secondary{background:#EEF2F7;color:#344054}.btn.warn{background:var(--accent);color:#111827}.btn.danger{background:var(--red)}.toolbar{display:flex;gap:10px;flex-wrap:wrap;align-items:center}.toolbar input,.toolbar select{height:39px;border:1px solid var(--line);border-radius:10px;background:white;padding:0 11px;min-width:150px}.toolbar input.search{min-width:280px}.tablebox{border:1px solid var(--line);border-radius:16px;overflow:hidden;background:white}.table-scroll{overflow:auto;max-height:68vh}table{border-collapse:separate;border-spacing:0;width:100%;font-size:13px}th{position:sticky;top:0;background:var(--nav);color:white;text-align:left;padding:10px 11px;white-space:nowrap;border-right:1px solid rgba(255,255,255,.17);z-index:2}td{padding:9px 11px;border-bottom:1px solid var(--line);border-right:1px solid #EEF2F7;vertical-align:top;max-width:480px}tr:nth-child(even) td{background:#FAFBFC}tr:hover td{background:#F1F7FF}.pagination{display:flex;justify-content:space-between;align-items:center;gap:12px;padding:12px 14px;border-top:1px solid var(--line);background:#FBFCFE}.muted{color:var(--muted)}.split{display:grid;grid-template-columns:1.15fr .85fr;gap:16px}.bars{display:grid;gap:11px}.barrow{display:grid;grid-template-columns:minmax(120px,1fr) 72px;gap:10px;align-items:center}.barrow span:first-child{white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-weight:700}.bar{height:10px;border-radius:99px;background:#EEF2F7;overflow:hidden;grid-column:1 / -1}.bar i{display:block;height:100%;background:linear-gradient(90deg,var(--blue),#7C3AED)}.formrow{display:grid;grid-template-columns:1fr 1fr auto;gap:12px;align-items:end}.field label{display:block;font-size:12px;font-weight:900;margin-bottom:5px}.field input{height:40px;width:100%;border:1px solid var(--line);border-radius:10px;padding:0 11px}.log{background:#0B1220;color:#D1D5DB;border-radius:16px;padding:16px;font-family:Consolas,monospace;white-space:pre-wrap;max-height:70vh;overflow:auto}.drawer{position:fixed;right:0;top:0;bottom:0;width:min(760px,92vw);background:white;border-left:1px solid var(--line);box-shadow:var(--shadow);transform:translateX(105%);transition:.2s;z-index:50;display:flex;flex-direction:column}.drawer.open{transform:translateX(0)}.drawer-h{padding:18px;border-bottom:1px solid var(--line);display:flex;align-items:center;justify-content:space-between}.drawer-b{padding:18px;overflow:auto}.detail{display:grid;grid-template-columns:240px 1fr;border:1px solid var(--line);border-bottom:0;border-radius:14px;overflow:hidden}.detail div{padding:9px 11px;border-bottom:1px solid var(--line)}.detail div:nth-child(odd){background:#F8FAFC;font-weight:900}.spinner{width:18px;height:18px;border:3px solid #D8E0EA;border-top-color:var(--blue);border-radius:50%;animation:spin .8s linear infinite}@keyframes spin{to{transform:rotate(360deg)}}.overlay{position:fixed;inset:0;background:rgba(11,18,32,.42);display:none;z-index:40}.overlay.open{display:block}.mobile{display:none}@media(max-width:1050px){.app{grid-template-columns:1fr}.side{position:fixed;z-index:60;transform:translateX(-105%);transition:.2s;width:280px}.side.open{transform:translateX(0)}.mobile{display:inline-flex}.cards{grid-template-columns:repeat(2,minmax(0,1fr))}.split,.formrow{grid-template-columns:1fr}.top{padding:0 14px}.content{padding:16px}.toolbar input.search{min-width:100%}}@media(max-width:620px){.cards{grid-template-columns:1fr}.toolbar input,.toolbar select,.btn{width:100%;min-width:100%}.detail{grid-template-columns:1fr}}
</style>
</head>
<body>
<div class="app">
  <aside class="side" id="side">
    <div class="brand"><h1>Motor Fiscal RIR70 by MC4</h1><p>MC4 CONTABILIDADE E GESTAO DE NEGOCIOS<br>CNPJ 09.944.432/0001-25<br>Ferramenta proprietaria licenciada</p></div>
    <nav class="nav">
      <div class="section">Navegacao</div>
      <button data-page="dashboard" class="active">Dashboard Executivo</button>
      <button data-page="executar">Executar Motor</button>
      <button data-page="capa">CAPA</button>
      <div class="section">Abas do Output</div>
      <button data-page="pendencias">Pendencias</button>
      <button data-page="arbitramento">Arbitramento</button>
      <button data-page="inventario">Inventario Valorizado</button>
      <button data-page="movimento">Movimento Validado</button>
      <button data-page="fontes">Fontes Processadas</button>
      <div class="section">Controle</div>
      <button data-page="integridade">Integridade MC4</button>
      <button data-page="log">Log</button>
    </nav>
  </aside>
  <main class="main">
    <header class="top">
      <button class="btn secondary mobile" onclick="toggleSide()">Menu</button>
      <h2 id="pageTitle">Dashboard Executivo</h2>
      <div class="top-actions"><span id="cachePill" class="pill muted">Cache</span><a class="btn secondary" href="/download_output">Baixar Excel Oficial</a></div>
    </header>
    <section class="content" id="content"></section>
  </main>
</div>
<div class="overlay" id="overlay" onclick="closeDrawer()"></div>
<aside class="drawer" id="drawer"><div class="drawer-h"><h3 id="drawerTitle">Detalhe</h3><button class="btn secondary" onclick="closeDrawer()">Fechar</button></div><div class="drawer-b" id="drawerBody"></div></aside>
<script>
const state = {page:'dashboard', sheet:null, pageNo:1, perPage:100, mode:'executivo', filters:{}, search:'', meta:null};
const sheetLabels = {pendencias:'Pendencias', arbitramento:'Arbitramento', inventario:'Inventario Valorizado', movimento:'Movimento Validado', fontes:'Fontes Processadas'};
function $(id){return document.getElementById(id)}
function esc(v){return (v??'').toString().replace(/[&<>"]/g, m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[m]))}
function fmt(n){return Number(n||0).toLocaleString('pt-BR')}
function setTitle(t){$('pageTitle').textContent=t}
function toggleSide(){ $('side').classList.toggle('open') }
function activate(page){document.querySelectorAll('.nav button').forEach(b=>b.classList.toggle('active', b.dataset.page===page)); if(innerWidth<1050)$('side').classList.remove('open')}
async function api(url, opts){const r=await fetch(url, opts); if(!r.ok) throw new Error(await r.text()); return await r.json()}
function loading(msg='Carregando...'){$('content').innerHTML=`<div class="card"><div class="card-b" style="display:flex;gap:12px;align-items:center"><span class="spinner"></span><strong>${esc(msg)}</strong></div></div>`}
async function ensureCache(){
  const st = await api('/api/cache/status'); updateCachePill(st);
  if(st.current) return true;
  $('content').innerHTML = `<div class="notice"><strong>Indexacao necessaria.</strong> O Excel atual ainda nao foi convertido para cache analitico. A navegacao ficara rapida apos esta etapa.</div><div class="card"><div class="card-h"><h3>Preparar sistema para navegacao</h3></div><div class="card-b"><p>O sistema usa cache SQLite local para evitar ler 200 mil linhas do Excel a cada clique.</p><button class="btn" onclick="startCacheBuild()">Indexar Excel agora</button><br><br><div id="cacheMsg" class="muted">${esc(st.job.message||'Aguardando')}</div></div></div>`;
  if(st.xlsx && st.xlsx.path) { startCacheBuild(); }
  return false;
}
async function startCacheBuild(){
  await api('/api/cache/start', {method:'POST'});
  pollCache();
}
async function pollCache(){
  const box=$('cacheMsg');
  const timer=setInterval(async()=>{
    const st=await api('/api/cache/status'); updateCachePill(st);
    if(box) box.innerHTML=`<span class="spinner"></span> ${esc(st.job.message||'Indexando')} ${st.job.sheet? ' - '+esc(st.job.sheet):''} | linhas: ${fmt(st.job.rows||0)}`;
    if(!st.job.running){ clearInterval(timer); if(st.current){ await loadPage(state.page); } else if(box){box.textContent='Falha ao indexar: '+(st.job.error||'erro desconhecido')} }
  }, 1300);
}
function updateCachePill(st){const p=$('cachePill'); if(!p)return; if(st.current){p.className='pill ok'; p.textContent='Cache OK'} else if(st.job&&st.job.running){p.className='pill rev'; p.textContent='Indexando'} else {p.className='pill muted'; p.textContent='Cache pendente'}}
async function loadPage(page){
  state.page=page; activate(page);
  if(['pendencias','arbitramento','inventario','movimento','fontes'].includes(page)){state.sheet=page; state.pageNo=1; state.filters={}; state.search=''; await renderSheet(page); return;}
  if(page==='dashboard'){await renderDashboard(); return}
  if(page==='executar'){renderExecutar(); return}
  if(page==='capa'){await renderCapa(); return}
  if(page==='integridade'){await renderIntegridade(); return}
  if(page==='log'){await renderLog(); return}
}
async function renderDashboard(){
  setTitle('Dashboard Executivo'); loading('Carregando dashboard');
  const ok = await ensureCache(); if(!ok)return;
  const d = await api('/api/overview'); updateCachePill(d.cache);
  const s=d.pendencias;
  $('content').innerHTML = `<div class="notice"><strong>${esc(d.brand)}</strong> - navegador = cache analitico derivado; Excel = evidencia fiscal oficial. Hash Excel: <code>${esc((d.cache.xlsx&&d.cache.xlsx.sha256)||'N/D')}</code></div>
  <div class="grid cards">
    <div class="card"><div class="card-b"><div class="metric crit">${fmt(s.critico)}<small>Pendencias criticas</small></div></div></div>
    <div class="card"><div class="card-b"><div class="metric rev">${fmt(s.revisar)}<small>Itens para revisar</small></div></div></div>
    <div class="card"><div class="card-b"><div class="metric blue">${fmt(s.total)}<small>Total de pendencias</small></div></div></div>
    <div class="card"><div class="card-b"><div class="metric ok">${esc(d.integridade.status)}<small>Integridade MC4</small></div></div></div>
  </div><br>
  <div class="split">
    <div class="card"><div class="card-h"><h3>Top tipos de pendencia</h3><button class="btn secondary" onclick="loadPage('pendencias')">Abrir Pendencias</button></div><div class="card-b">${bars(s.por_tipo)}</div></div>
    <div class="card"><div class="card-h"><h3>Abas do output</h3></div><div class="card-b"><div class="bars">${Object.values(d.meta.sheets).map(x=>`<div class="barrow"><span>${esc(x.label)}</span><strong>${fmt(x.row_count)}</strong><div class="bar"><i style="width:${Math.min(100, (x.row_count/Math.max(1,d.max_rows))*100)}%"></i></div></div>`).join('')}</div></div></div>
  </div><br>
  <div class="card"><div class="card-h"><h3>Output oficial indexado</h3><button class="btn secondary" onclick="startCacheBuild()">Reindexar Excel</button></div><div class="card-b"><strong>${esc(d.cache.xlsx.name||'Nenhum arquivo')}</strong><br><span class="muted">${esc(d.cache.xlsx.path||'')}</span><br><small>SHA-256 Excel: ${esc(d.cache.xlsx.sha256||'N/D')}</small><br><small>SHA-256 Cache: ${esc((d.cache.state&&d.cache.state.cache_sha256)||'N/D')}</small><br><small>${esc((d.cache.state&&d.cache.state.aviso)||'')}</small></div></div>`;
}
function bars(rows){ if(!rows||!rows.length) return '<span class="muted">Sem dados.</span>'; const max=Math.max(...rows.map(r=>r.value||0),1); return `<div class="bars">${rows.map(r=>`<div class="barrow"><span title="${esc(r.label)}">${esc(r.label)}</span><strong>${fmt(r.value)}</strong><div class="bar"><i style="width:${Math.max(3,(r.value/max)*100)}%"></i></div></div>`).join('')}</div>` }
async function renderSheet(sheet){
  setTitle(sheetLabels[sheet]||sheet); loading('Carregando '+(sheetLabels[sheet]||sheet)); const ok=await ensureCache(); if(!ok)return; await fetchSheet();
}
function sheetShell(data){
  const ds=data.distinct||{};
  return `<div class="notice"><strong>${esc(data.label)}</strong> - ${esc(data.help||'Use filtros, busca e detalhe por linha para analisar.')}</div>
  <div class="card"><div class="card-h"><h3>Filtros e navegacao</h3><div><span class="pill muted">${fmt(data.total)} registros</span></div></div><div class="card-b">
    <div class="toolbar">
      <input class="search" id="searchBox" placeholder="Buscar por produto, CNPJ, NCM, CFOP, chave..." value="${esc(state.search)}" onkeydown="if(event.key==='Enter'){state.search=this.value;state.pageNo=1;fetchSheet()}">
      ${selectHtml('empresa','Empresa',ds.empresa)}${selectHtml('mes','Mes',ds.mes)}${selectHtml('nivel','Nivel/Status',ds.nivel)}${selectHtml('tipo','Tipo',ds.tipo)}
      <select id="modeSel" onchange="state.mode=this.value;state.pageNo=1;fetchSheet()"><option value="executivo" ${state.mode==='executivo'?'selected':''}>Visao executiva</option><option value="compacto" ${state.mode==='compacto'?'selected':''}>Compacta</option><option value="todas" ${state.mode==='todas'?'selected':''}>Todas as colunas</option></select>
      <button class="btn" onclick="state.search=$('searchBox').value;state.pageNo=1;fetchSheet()">Filtrar</button><button class="btn secondary" onclick="clearFilters()">Limpar</button><a class="btn secondary" href="/api/export/${state.sheet}?${queryParams(true)}">CSV</a>
    </div>
  </div></div><br><div class="tablebox"><div class="table-scroll"><table><thead><tr>${data.headers.map(h=>`<th>${esc(h)}</th>`).join('')}<th>Detalhe</th></tr></thead><tbody>${data.rows.map(r=>rowHtml(r,data.headers)).join('')||`<tr><td colspan="${data.headers.length+1}">Sem dados.</td></tr>`}</tbody></table></div><div class="pagination"><span>Pagina ${fmt(data.page)} de ${fmt(data.pages)} | ${fmt(data.total)} registros</span><div class="toolbar"><button class="btn secondary" ${data.page<=1?'disabled':''} onclick="state.pageNo--;fetchSheet()">Anterior</button><select onchange="state.perPage=this.value;state.pageNo=1;fetchSheet()"><option ${state.perPage==50?'selected':''}>50</option><option ${state.perPage==100?'selected':''}>100</option><option ${state.perPage==250?'selected':''}>250</option><option ${state.perPage==500?'selected':''}>500</option></select><button class="btn secondary" ${data.page>=data.pages?'disabled':''} onclick="state.pageNo++;fetchSheet()">Proxima</button></div></div></div>`;
}
function rowHtml(r,headers){return `<tr>${headers.map(h=>`<td>${cellHtml(h,r[h])}</td>`).join('')}<td><button class="btn secondary" onclick="openDetail('${state.sheet}',${r.rid})">Abrir</button></td></tr>`}
function cellHtml(h,v){ const t=(v??'').toString(); const up=t.toUpperCase(); if(h.toUpperCase().includes('NIVEL')||h.toUpperCase().includes('STATUS')||h.toUpperCase().includes('SITUACAO')){let cls=up.includes('CRITICO')?'crit':(up.includes('REVISAR')||up.includes('PENDENTE')?'rev':(up.includes('OK')||up.includes('APROV')?'ok':'muted')); return `<span class="pill ${cls}">${esc(t)}</span>`} return esc(t)}
function selectHtml(key,label,items){return `<select id="f_${key}" onchange="state.filters['${key}']=this.value;state.pageNo=1;fetchSheet()"><option value="">${label}</option>${(items||[]).map(x=>`<option value="${esc(x)}" ${(state.filters[key]||'')===x?'selected':''}>${esc(x)}</option>`).join('')}</select>`}
function queryParams(exporting=false){const p=new URLSearchParams(); p.set('page',state.pageNo); p.set('per_page', exporting?50000:state.perPage); p.set('mode',state.mode); if(state.search)p.set('q',state.search); for(const [k,v] of Object.entries(state.filters)){if(v)p.set(k,v)} return p.toString()}
async function fetchSheet(){loading('Consultando cache'); const data=await api(`/api/sheet/${state.sheet}?${queryParams()}`); $('content').innerHTML=sheetShell(data)}
function clearFilters(){state.filters={};state.search='';state.pageNo=1;fetchSheet()}
async function openDetail(sheet,rid){const d=await api(`/api/row/${sheet}/${rid}`); $('drawerTitle').textContent=`Detalhe - linha ${rid}`; $('drawerBody').innerHTML=`<div class="detail">${Object.entries(d).filter(([k])=>k!=='rid').map(([k,v])=>`<div>${esc(k)}</div><div>${esc(v)}</div>`).join('')}</div>`; $('overlay').classList.add('open'); $('drawer').classList.add('open')}
function closeDrawer(){ $('overlay').classList.remove('open'); $('drawer').classList.remove('open') }
async function renderCapa(){ setTitle('CAPA'); state.sheet='capa'; state.mode='todas'; state.pageNo=1; await renderSheet('capa') }
function renderExecutar(){ setTitle('Executar Motor'); $('content').innerHTML=`<div class="notice"><strong>Execucao controlada.</strong> Informe o periodo e acompanhe o status. Ao final, reindexe o Excel para navegar rapidamente.</div><div class="card"><div class="card-h"><h3>Periodo de processamento</h3><span id="runPill" class="pill muted">Aguardando</span></div><div class="card-b"><div class="formrow"><div class="field"><label>Data inicial (DD/MM/AAAA)</label><input id="dtIni" placeholder="01/01/2026"></div><div class="field"><label>Data final (DD/MM/AAAA)</label><input id="dtFim" placeholder="31/12/2026"></div><button class="btn" onclick="runMotor()">Executar motor</button></div><br><div id="runMsg" class="muted">Pronto.</div></div></div><br><div class="card"><div class="card-h"><h3>Saida do motor</h3><button class="btn secondary" onclick="pollRun(true)">Atualizar</button></div><div class="card-b"><div class="log" id="runLog">Sem log nesta sessao.</div></div></div>`; pollRun(false)}
async function runMotor(){const di=$('dtIni').value, df=$('dtFim').value; await api('/api/run',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({data_inicio:di,data_fim:df})}); pollRun(true)}
async function pollRun(loop){const r=await api('/api/run_status'); const p=$('runPill'); if(p){p.className='pill '+(r.running?'rev':(r.exitcode===0?'ok':(r.exitcode?'crit':'muted'))); p.textContent=r.running?'Executando':(r.exitcode===0?'Concluido':(r.exitcode?'Erro':'Aguardando'))} if($('runMsg'))$('runMsg').textContent=r.message||''; if($('runLog'))$('runLog').textContent=(r.tail||[]).join('\n')||'Sem log nesta sessao.'; if(loop&&r.running)setTimeout(()=>pollRun(true),1500)}
async function renderIntegridade(){setTitle('Integridade MC4'); loading(); const d=await api('/api/integridade'); $('content').innerHTML=`<div class="notice"><strong>Controle proprietario MC4.</strong> Verifica hashes dos arquivos criticos do pacote.</div><div class="grid cards"><div class="card"><div class="card-b"><div class="metric ${d.status==='OK'?'ok':'crit'}">${esc(d.status)}<small>Status</small></div></div></div><div class="card"><div class="card-b"><div class="metric blue">${fmt(d.arquivos.length)}<small>Arquivos criticos</small></div></div></div><div class="card"><div class="card-b"><div class="metric crit">${fmt(d.divergentes.length)}<small>Divergencias</small></div></div></div><div class="card"><div class="card-b"><div class="metric">${esc(d.manifesto.versao||'N/D')}<small>Versao</small></div></div></div></div><br><div class="tablebox"><div class="table-scroll"><table><thead><tr><th>Arquivo</th><th>Status</th><th>Esperado</th><th>Encontrado</th></tr></thead><tbody>${d.arquivos.map(r=>`<tr><td>${esc(r.arquivo)}</td><td><span class="pill ${r.status==='OK'?'ok':'crit'}">${esc(r.status)}</span></td><td>${esc(r.esperado)}</td><td>${esc(r.encontrado)}</td></tr>`).join('')}</tbody></table></div></div>`}
async function renderLog(){setTitle('Log'); loading(); const d=await api('/api/log'); $('content').innerHTML=`<div class="card"><div class="card-h"><h3>Ultimas linhas do log</h3><button class="btn secondary" onclick="renderLog()">Atualizar</button></div><div class="card-b"><div class="log">${esc((d.lines||[]).join('\n'))}</div></div></div>`}

document.querySelectorAll('.nav button').forEach(b=>b.addEventListener('click',()=>loadPage(b.dataset.page)));
loadPage('dashboard');
</script>
</body>
</html>
"""


@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/api/cache/status")
def api_cache_status():
    return jsonify(cache_status_payload())


@app.route("/api/cache/start", methods=["POST"])
def api_cache_start():
    force = (request.args.get("force") == "1")
    return jsonify(start_cache_build(force=force))


@app.route("/api/overview")
def api_overview():
    if not cache_current():
        return jsonify({"cache": cache_status_payload(), "meta": get_meta(), "pendencias": load_summary("pendencias"), "integridade": integrity_summary(), "brand": MOTOR_BRAND, "max_rows": 1})
    meta = get_meta()
    max_rows = max([s.get("row_count", 0) for s in meta.get("sheets", {}).values()] or [1])
    return jsonify({"cache": cache_status_payload(), "meta": meta, "pendencias": load_summary("pendencias"), "integridade": integrity_summary(), "brand": MOTOR_BRAND, "max_rows": max_rows})


@app.route("/api/sheet/<sheet_key>")
def api_sheet(sheet_key: str):
    if sheet_key not in SHEET_MAP:
        return jsonify({"error": "aba invalida"}), 404
    filters = {k: request.args.get(k, "") for k in ["empresa", "mes", "nivel", "tipo"]}
    return jsonify(sheet_query(sheet_key, int(request.args.get("page", 1)), int(request.args.get("per_page", 100)), request.args.get("q", ""), filters, request.args.get("mode", "executivo")))


@app.route("/api/row/<sheet_key>/<int:rid>")
def api_row(sheet_key: str, rid: int):
    if sheet_key not in SHEET_MAP:
        return jsonify({"error": "aba invalida"}), 404
    return jsonify(read_row(sheet_key, rid))


@app.route("/api/export/<sheet_key>")
def api_export(sheet_key: str):
    if sheet_key not in SHEET_MAP:
        return "aba invalida", 404
    filters = {k: request.args.get(k, "") for k in ["empresa", "mes", "nivel", "tipo"]}
    data = sheet_query(sheet_key, 1, int(request.args.get("per_page", 50000)), request.args.get("q", ""), filters, request.args.get("mode", "executivo"))
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(data.get("headers", []))
    for row in data.get("rows", []):
        writer.writerow([row.get(h, "") for h in data.get("headers", [])])
    return Response(output.getvalue(), mimetype="text/csv; charset=utf-8", headers={"Content-Disposition": f"attachment; filename={sheet_key}_filtrado.csv"})


@app.route("/api/integridade")
def api_integridade():
    return jsonify(integrity_summary())


@app.route("/api/log")
def api_log():
    return jsonify({"lines": load_log_lines()})


@app.route("/api/run", methods=["POST"])
def api_run():
    if _run_status.get("running"):
        return jsonify(_run_status)
    data = request.get_json(silent=True) or {}
    di = safe_text(data.get("data_inicio"))
    df = safe_text(data.get("data_fim"))
    if not di or not df:
        return jsonify({"error": "Informe data_inicio e data_fim"}), 400
    _run_env_vars.clear()
    _run_env_vars.update({"RIR70_DATA_INICIAL": di, "RIR70_DATA_FINAL": df})
    _last_form.update({"data_inicio": di, "data_fim": df})
    threading.Thread(target=run_motor_thread, daemon=True).start()
    return jsonify(_run_status)


@app.route("/api/run_status")
def api_run_status():
    return jsonify({**_run_status, "tail": _run_output_tail[-300:]})


@app.route("/download_output")
def download_output():
    xlsx = last_output_path()
    if not xlsx or not xlsx.exists():
        return "Nenhum output encontrado", 404
    return send_file(xlsx, as_attachment=True, download_name=xlsx.name)


@app.route("/health")
def health():
    return jsonify({"status": "OK", "version": APP_VERSION, "brand": MOTOR_BRAND})


if __name__ == "__main__":
    print("=" * 62)
    print("MC4 CONTABILIDADE E GESTAO DE NEGOCIOS")
    print("CNPJ 09.944.432/0001-25")
    print("Motor Fiscal RIR70 by MC4 - Sistema Web Performance")
    print("Acesse: http://127.0.0.1:5000")
    print("=" * 62)
    app.run(host="127.0.0.1", port=5000, debug=False, threaded=True)
